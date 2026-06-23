# -*- coding: utf-8 -*-
"""
Análisis automático de transcripciones GNS - Banco Atlántida HN
Genera un Excel con:
  1. Base Clasificada
  2. Pareto
  3. Resumen Ejecutivo

Columnas del CSV de entrada (GNS_Transcripciones_YYYYMMDD_HHMMSS.csv):
  conversationId, communicationId, conversationStart, conversationEnd,
  originatingDirection, ContactId, ContactListId, CampaignId,
  phrases_count, transcript_first_marker, transcript_last_marker,
  text, fecha_carga

Dependencias:
  pip install pandas openpyxl

Parámetros esperados:
  INPUT_FILE_PATH  = Ruta completa del CSV o Excel de transcripciones
  OUTPUT_DIR       = Carpeta donde se guardará el Excel final
  OUTPUT_FILENAME  = Nombre opcional del archivo final
"""

import os
import re
import sys
import argparse
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PARÁMETROS PARA PYFLOW MANAGER
# ============================================================
PYFLOW_PARAMS = {
    "INPUT_FILE_PATH": {
        "type": "text",
        "label": "Ruta del archivo de transcripciones",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports\GNS_Transcripciones_YYYYMMDD_HHMMSS.csv"
    },
    "OUTPUT_DIR": {
        "type": "text",
        "label": "Ruta de salida",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports"
    },
    "OUTPUT_FILENAME": {
        "type": "text",
        "label": "Nombre del archivo final",
        "required": False,
        "default": ""
    }
}


# ============================================================
# COLUMNAS DEL CSV GNS (en el orden que aparecen en la Base Clasificada)
# ============================================================
# Columnas originales que SÍ se incluyen en el output (se excluyen las vacías)
COLUMNAS_INCLUIR = [
    "conversationId",
    "communicationId",
    "conversationStart",
    "conversationEnd",
    "originatingDirection",
    "phrases_count",
    "fecha_carga",
]
# Columnas que se descartan por estar vacías en la fuente
COLUMNAS_DESCARTAR = [
    "ContactId", "ContactListId", "CampaignId",
    "transcript_first_marker", "transcript_last_marker"
]


# ============================================================
# UTILIDADES
# ============================================================
def normalizar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def contiene(texto: str, patrones) -> bool:
    return any(re.search(p, texto, flags=re.IGNORECASE) for p in patrones)


def calcular_duracion_segundos(row) -> float:
    """Calcula duración de la llamada en segundos desde conversationStart/End."""
    try:
        inicio = pd.to_datetime(row["conversationStart"], utc=True)
        fin = pd.to_datetime(row["conversationEnd"], utc=True)
        return round((fin - inicio).total_seconds(), 0)
    except Exception:
        return None


def segundos_a_mmss(segundos) -> str:
    """Convierte segundos a formato MM:SS."""
    if pd.isna(segundos) or segundos is None:
        return ""
    s = int(segundos)
    return f"{s // 60:02d}:{s % 60:02d}"


def leer_archivo(ruta: str) -> pd.DataFrame:
    ruta = Path(ruta)

    if not ruta.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {ruta}")

    if ruta.suffix.lower() == ".csv":
        try:
            return pd.read_csv(ruta, encoding="utf-8-sig")
        except UnicodeDecodeError:
            return pd.read_csv(ruta, encoding="latin-1")

    if ruta.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(ruta)

    raise ValueError("Formato no soportado. Use CSV, XLSX o XLS.")


# ============================================================
# CLASIFICACIÓN PROFUNDA POR MOTIVO RAÍZ
# ============================================================
def extraer_evidencia(texto: str, patrones) -> str:
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            inicio = max(0, m.start() - 70)
            fin = min(len(texto), m.end() + 90)
            return texto[inicio:fin].strip()
    return ""


def clasificar_motivo(transcripcion) -> dict:
    t = normalizar_texto(transcripcion)

    if not t or len(t) < 20:
        return {
            "Motivo raíz": "No identificado / requiere revisión",
            "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja"
        }

    reglas = [
        # --- Orden de prioridad: de más específico a más genérico ---
        {
            "motivo": "Cambio de dispositivo / reinstalación",
            "submotivo": "El cliente cambió celular, reinstaló la app o perdió acceso desde otro equipo",
            "patrones": [
                r"cambie.*(celular|telefono|equipo|dispositivo)", r"nuevo celular",
                r"nuevo telefono", r"reinstal", r"instale de nuevo", r"otro dispositivo",
                r"perdi.*celular", r"formatee", r"desinstale"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Contraseña temporal vencida / recuperación incompleta",
            "submotivo": "Clave temporal o enlace de recuperación expiró antes de completar el proceso",
            "patrones": [
                r"temporal.*vencid", r"contrasena.*temporal.*expir", r"clave.*temporal.*caduc",
                r"enlace.*venc", r"recuperacion incompleta", r"temporal.*caduc"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Migración a Atlántida HN / nueva app",
            "submotivo": "Credenciales o flujo anterior no funcionan en nueva plataforma Atlántida HN",
            "patrones": [
                r"atlantida hn", r"atlantia", r"nueva app", r"nueva aplicacion",
                r"nueva plataforma", r"migracion", r"migrar", r"banca digital nueva",
                r"actualizar la app", r"app nueva", r"antes ingresaba",
                r"credenciales invalidas", r"credenciales de autenticacion"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Usuario bloqueado / requiere desbloqueo",
            "submotivo": "El usuario está bloqueado o el cliente solicita desbloqueo",
            "patrones": [
                r"bloquead", r"desbloque", r"usuario bloqueado", r"me bloqueo",
                r"se bloqueo", r"desbloquear usuario", r"cuenta bloqueada"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe código OTP / SMS / Token",
            "submotivo": "El código de verificación no llega o falla el token",
            "patrones": [
                r"no.*(llega|recibo|manda|envia).*(codigo|otp|token|sms|mensaje|correo)",
                r"(codigo|otp|token|sms).*(no.*llega|no.*recibo|no.*manda|no.*envia)",
                r"codigo de verificacion", r"no me cae el codigo",
                r"no me llega.*correo", r"correo.*codigo"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Actualización de datos / teléfono o correo asociado",
            "submotivo": "Datos registrados no permiten validar identidad o recuperar acceso",
            "patrones": [
                r"actualizar.*(datos|telefono|correo|numero)",
                r"cambiar.*(telefono|correo|numero)",
                r"telefono.*registrado", r"correo.*registrado", r"numero.*registrado",
                r"ya no tengo.*(numero|telefono|correo)", r"otro numero", r"otro correo",
                r"datos desactualizados", r"actualizacion de datos", r"validar identidad",
                r"aboque.*agencia", r"agencia.*actualiz"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe correo / correo incorrecto",
            "submotivo": "El enlace, correo o notificación de recuperación no llega o el correo es incorrecto",
            "patrones": [
                r"correo incorrecto", r"correo malo", r"no recibe correo",
                r"no me llega.*correo", r"email.*no llega", r"bandeja.*spam",
                r"correo electronico"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Cliente nuevo / primer acceso",
            "submotivo": "Cliente recién afiliado o intentando entrar por primera vez",
            "patrones": [
                r"primera vez", r"primer acceso", r"cliente nuevo", r"recien afiliad",
                r"me afilie", r"afiliacion", r"activar usuario", r"crear usuario",
                r"nunca he entrado", r"entrar por primera"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Dificultad tecnológica / requiere guía",
            "submotivo": "El cliente necesita acompañamiento paso a paso para completar el proceso",
            "patrones": [
                r"no se como", r"ayudeme", r"guieme", r"paso a paso",
                r"no entiendo", r"me cuesta", r"no puedo hacerlo", r"no manejo"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Olvidó contraseña / no recuerda credenciales",
            "submotivo": "El cliente manifiesta olvido de clave, contraseña o usuario",
            "patrones": [
                r"olvide", r"olvido", r"no recuerdo", r"no me acuerdo",
                r"recuperar.*(clave|contrasena|usuario)",
                r"cambiar.*(contrasena|clave)", r"restablecer.*(contrasena|clave)",
                r"resetear.*(contrasena|clave)", r"contrasena temporal",
                r"contrasena", r"contraseña", r"clave"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Error de app/web o autenticación",
            "submotivo": "La app/web muestra error, no permite avanzar o rechaza credenciales",
            "patrones": [
                r"error", r"no funciona", r"falla", r"problema con la app",
                r"no me deja", r"no permite", r"rechaza", r"autenticacion",
                r"pantalla", r"se queda cargando", r"usuario invalido",
                r"no puedo entrar", r"no puedo ingresar"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Consulta operativa / validación general NBDA",
            "submotivo": "La llamada no es claramente cambio de contraseña; es soporte general de banca digital",
            "patrones": [
                r"consulta", r"informacion", r"ayuda", r"soporte", r"banca digital",
                r"transferencia", r"pago", r"saldo", r"movimiento", r"cuenta"
            ],
            "confianza": "Media"
        },
    ]

    for regla in reglas:
        if contiene(t, regla["patrones"]):
            evidencia = extraer_evidencia(t, regla["patrones"])
            return {
                "Motivo raíz": regla["motivo"],
                "Submotivo raíz": regla["submotivo"],
                "Evidencia detectada": evidencia,
                "Nivel de confianza": regla["confianza"]
            }

    return {
        "Motivo raíz": "No identificado / requiere revisión",
        "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
        "Evidencia detectada": "",
        "Nivel de confianza": "Baja"
    }


# ============================================================
# PREPARACIÓN DE HOJAS
# ============================================================
def preparar_base_clasificada(df: pd.DataFrame) -> pd.DataFrame:
    """
    Construye la hoja Base Clasificada con:
      - 4 columnas de clasificación (al frente)
      - Columnas originales útiles del CSV
      - Duración en segundos y MM:SS
      - Transcripción completa al final
    """
    clasificaciones = df["text"].apply(clasificar_motivo).apply(pd.Series)

    # Calcular duración
    df = df.copy()
    df["Duracion_seg"] = df.apply(calcular_duracion_segundos, axis=1)
    df["Duración (MM:SS)"] = df["Duracion_seg"].apply(segundos_a_mmss)

    # Columnas que sí existen en el CSV (filtramos las vacías)
    cols_base = [c for c in COLUMNAS_INCLUIR if c in df.columns]

    base = pd.DataFrame()
    # 1. Clasificación
    base["Motivo raíz"] = clasificaciones["Motivo raíz"]
    base["Submotivo raíz"] = clasificaciones["Submotivo raíz"]
    base["Evidencia detectada"] = clasificaciones["Evidencia detectada"]
    base["Nivel de confianza"] = clasificaciones["Nivel de confianza"]
    # 2. Datos de la llamada
    for col in cols_base:
        base[col] = df[col].values
    base["Duración (MM:SS)"] = df["Duración (MM:SS)"].values
    base["phrases_count"] = df["phrases_count"].values
    # 3. Transcripción al final
    base["Transcripción completa"] = df["text"].values

    # Evitar duplicado de phrases_count si ya estaba en cols_base
    base = base.loc[:, ~base.columns.duplicated()]

    return base


def preparar_pareto(base: pd.DataFrame) -> pd.DataFrame:
    pareto = (
        base.groupby(["Motivo raíz", "Submotivo raíz"], dropna=False)
        .size()
        .reset_index(name="Llamadas")
        .sort_values("Llamadas", ascending=False)
        .reset_index(drop=True)
    )

    total = pareto["Llamadas"].sum()
    pareto["%"] = pareto["Llamadas"] / total
    pareto["% acumulado"] = pareto["%"].cumsum()

    total_row = pd.DataFrame([{
        "Motivo raíz": "Total general",
        "Submotivo raíz": "",
        "Llamadas": total,
        "%": 1.0,
        "% acumulado": 1.0
    }])

    return pd.concat([pareto, total_row], ignore_index=True)


def preparar_resumen_ejecutivo(df_original: pd.DataFrame, base: pd.DataFrame, pareto: pd.DataFrame) -> list:
    """
    Devuelve lista de (etiqueta, valor) para la hoja Resumen Ejecutivo.
    """
    total = len(base)
    inbound = (df_original["originatingDirection"] == "inbound").sum()
    outbound = (df_original["originatingDirection"] == "outbound").sum()

    duraciones = base["Duración (MM:SS)"].apply(
        lambda x: int(x.split(":")[0]) * 60 + int(x.split(":")[1]) if x else None
    ).dropna()
    dur_prom = segundos_a_mmss(duraciones.mean()) if len(duraciones) else "N/D"
    dur_max = segundos_a_mmss(duraciones.max()) if len(duraciones) else "N/D"
    dur_min = segundos_a_mmss(duraciones.min()) if len(duraciones) else "N/D"

    conf_alta = (base["Nivel de confianza"] == "Alta").sum()
    conf_media = (base["Nivel de confianza"] == "Media").sum()
    conf_baja = (base["Nivel de confianza"] == "Baja").sum()

    motivo_top = pareto[pareto["Motivo raíz"] != "Total general"].iloc[0]

    fecha_inicio = pd.to_datetime(df_original["conversationStart"]).min()
    fecha_fin = pd.to_datetime(df_original["conversationStart"]).max()

    rows = [
        ("RESUMEN EJECUTIVO", ""),
        ("", ""),
        ("Período analizado", ""),
        ("  Fecha más antigua", str(fecha_inicio.date()) if not pd.isna(fecha_inicio) else "N/D"),
        ("  Fecha más reciente", str(fecha_fin.date()) if not pd.isna(fecha_fin) else "N/D"),
        ("", ""),
        ("VOLUMEN DE LLAMADAS", ""),
        ("  Total llamadas", total),
        ("  Inbound", int(inbound)),
        ("  Outbound", int(outbound)),
        ("", ""),
        ("DURACIÓN DE LLAMADAS", ""),
        ("  Promedio (MM:SS)", dur_prom),
        ("  Máximo (MM:SS)", dur_max),
        ("  Mínimo (MM:SS)", dur_min),
        ("", ""),
        ("CLASIFICACIÓN", ""),
        ("  Motivo raíz más frecuente", motivo_top["Motivo raíz"]),
        ("  Llamadas en motivo top", int(motivo_top["Llamadas"])),
        ("  % del total", f"{motivo_top['%']:.1%}"),
        ("", ""),
        ("NIVEL DE CONFIANZA", ""),
        ("  Alta", int(conf_alta)),
        ("  Media", int(conf_media)),
        ("  Baja (requiere revisión)", int(conf_baja)),
    ]
    return rows


# ============================================================
# FORMATO EXCEL
# ============================================================
VERDE_HEADER = "215E4B"
ROJO_HEADER = "E31B23"
AZUL_HEADER = "1F4E79"
BLANCO = "FFFFFF"
AZUL_CLARO = "D9EAF7"
GRIS_CLARO = "F2F2F2"
BORDE_COLOR = "B7C9E2"


def _borde():
    thin = Side(style="thin", color=BORDE_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color):
    return PatternFill("solid", fgColor=color)


def _header_font():
    return Font(color=BLANCO, bold=True, name="Arial", size=10)


def _data_font(bold=False):
    return Font(name="Arial", size=9, bold=bold)


def formato_base_clasificada(ws):
    border = _borde()

    # Encabezados
    for cell in ws[1]:
        cell.fill = _header_fill(VERDE_HEADER)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Datos
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill_row = PatternFill("solid", fgColor=GRIS_CLARO) if i % 2 == 0 else None
        for cell in row:
            cell.border = border
            cell.font = _data_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill_row:
                cell.fill = fill_row

    # Anchos de columna
    anchos = {
        "Motivo raíz": 42,
        "Submotivo raíz": 60,
        "Evidencia detectada": 70,
        "Nivel de confianza": 18,
        "conversationId": 38,
        "communicationId": 38,
        "conversationStart": 22,
        "conversationEnd": 22,
        "originatingDirection": 18,
        "phrases_count": 14,
        "Duración (MM:SS)": 16,
        "fecha_carga": 20,
        "Transcripción completa": 90,
    }
    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        width = anchos.get(header, max(14, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def formato_pareto(ws):
    border = _borde()

    # Encabezados
    for cell in ws[1]:
        cell.fill = _header_fill(ROJO_HEADER)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Datos
    for row in ws.iter_rows(min_row=2):
        motivo = ws.cell(row=row[0].row, column=1).value

        if motivo == "Total general":
            for cell in row:
                cell.fill = _header_fill(ROJO_HEADER)
                cell.font = Font(color=BLANCO, bold=True, name="Arial", size=9)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            for cell in row:
                cell.border = border
                cell.font = _data_font()
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Resaltar motivo raíz
            row[0].font = Font(bold=True, name="Arial", size=9)
            row[0].fill = _header_fill(AZUL_CLARO)

        # Formato porcentaje columnas 4 y 5
        ws.cell(row=row[0].row, column=4).number_format = "0.00%"
        ws.cell(row=row[0].row, column=5).number_format = "0.00%"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def formato_resumen(ws, rows_data):
    border = _borde()

    for i, (etiqueta, valor) in enumerate(rows_data, start=1):
        c_etiq = ws.cell(row=i, column=1, value=etiqueta)
        c_val = ws.cell(row=i, column=2, value=valor)

        if etiqueta == "RESUMEN EJECUTIVO":
            for c in [c_etiq, c_val]:
                c.fill = _header_fill(AZUL_HEADER)
                c.font = Font(color=BLANCO, bold=True, name="Arial", size=12)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
        elif etiqueta in ("VOLUMEN DE LLAMADAS", "DURACIÓN DE LLAMADAS",
                          "CLASIFICACIÓN", "NIVEL DE CONFIANZA", "Período analizado"):
            for c in [c_etiq, c_val]:
                c.fill = _header_fill(VERDE_HEADER)
                c.font = Font(color=BLANCO, bold=True, name="Arial", size=10)
                c.alignment = Alignment(vertical="center")
                c.border = border
        elif etiqueta == "":
            pass
        else:
            c_etiq.font = _data_font(bold=True)
            c_val.font = _data_font()
            c_etiq.border = border
            c_val.border = border
            c_etiq.alignment = Alignment(vertical="center")
            c_val.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A2"


def aplicar_formato_excel(output_path: str, resumen_rows: list):
    wb = load_workbook(output_path)

    if "Base Clasificada" in wb.sheetnames:
        formato_base_clasificada(wb["Base Clasificada"])

    if "Pareto" in wb.sheetnames:
        formato_pareto(wb["Pareto"])

    if "Resumen Ejecutivo" in wb.sheetnames:
        formato_resumen(wb["Resumen Ejecutivo"], resumen_rows)

    wb.save(output_path)


# ============================================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================================
def generar_excel(input_file_path: str, output_dir: str, output_filename: str = "") -> str:
    print(f"[1/5] Leyendo archivo: {input_file_path}")
    df = leer_archivo(input_file_path)
    print(f"      {len(df)} registros cargados | Columnas: {df.columns.tolist()}")

    print("[2/5] Clasificando transcripciones...")
    base = preparar_base_clasificada(df)

    print("[3/5] Generando Pareto y Resumen...")
    pareto = preparar_pareto(base)
    resumen_rows = preparar_resumen_ejecutivo(df, base, pareto)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not output_filename:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Analisis_Profundo_Transcripciones_{fecha}.xlsx"

    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    output_path = output_dir / output_filename

    print(f"[4/5] Escribiendo Excel: {output_path}")
    resumen_df = pd.DataFrame(resumen_rows, columns=["Métrica", "Valor"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Base Clasificada", index=False)
        pareto.to_excel(writer, sheet_name="Pareto", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen Ejecutivo", index=False, header=False)

    print("[5/5] Aplicando formato...")
    aplicar_formato_excel(str(output_path), resumen_rows)

    print(f"\n✅ Archivo generado: {output_path}")
    return str(output_path)


# ============================================================
# ENTRADA PRINCIPAL
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Análisis profundo de transcripciones GNS - Banco Atlántida HN"
    )
    parser.add_argument("--INPUT_FILE_PATH", default=os.getenv("INPUT_FILE_PATH", ""))
    parser.add_argument("--OUTPUT_DIR", default=os.getenv("OUTPUT_DIR", ""))
    parser.add_argument("--OUTPUT_FILENAME", default=os.getenv("OUTPUT_FILENAME", ""))

    args = parser.parse_args()

    if not args.INPUT_FILE_PATH:
        raise ValueError("Debe indicar INPUT_FILE_PATH.")
    if not args.OUTPUT_DIR:
        raise ValueError("Debe indicar OUTPUT_DIR.")

    generar_excel(
        input_file_path=args.INPUT_FILE_PATH,
        output_dir=args.OUTPUT_DIR,
        output_filename=args.OUTPUT_FILENAME
    )


if __name__ == "__main__":
    main()
