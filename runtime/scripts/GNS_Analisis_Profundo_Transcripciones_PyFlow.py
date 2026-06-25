# =========================================================
# GNS ANALISIS PROFUNDO DE TRANSCRIPCIONES - CSV/Excel -> Excel
# =========================================================
#
# Objetivo:
#   Clasificar transcripciones de llamadas por motivo raíz y generar:
#     1) Base Clasificada
#     2) Pareto
#
# Diseñado para importarse en PyFlow Manager.
#
# Dependencias:
#   pip install pandas openpyxl
#
# Parámetros:
#   INPUT_FILE_PATH : Ruta completa del CSV o Excel generado por el extractor.
#   OUTPUT_DIR      : Carpeta donde se guardará el Excel final.
#   OUTPUT_FILENAME : Nombre opcional del archivo final.
# =========================================================

import os
import re
import sys
import argparse
import logging
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


PYFLOW_PARAMS = {
    "INPUT_FILE_PATH": {"type": "text", "label": "Ruta completa del archivo CSV/Excel de transcripciones", "required": True},
    "OUTPUT_DIR": {"type": "text", "label": "Ruta de salida del Excel", "required": True},
    "OUTPUT_FILENAME": {"type": "text", "label": "Nombre del archivo Excel final opcional", "required": False}
}


LOGGER_NAME = "gns_analisis_profundo_transcripciones"


@dataclass
class Config:
    input_file_path: str
    output_dir: str
    output_filename: str


def setup_logger() -> logging.Logger:
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        "%Y-%m-%d %H:%M:%S"
    )

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(formatter)
    logger.addHandler(console)

    return logger


def _clean_env_value(value: Any, default: Optional[str] = None) -> Optional[str]:
    if value is None:
        return default
    text = str(value).strip()
    if text == "" or text.lower() in ("null", "none", "undefined"):
        return default
    return text


def env_str(name: str, default: Optional[str] = None, required: bool = False) -> str:
    value = _clean_env_value(os.getenv(name), default)
    if required and not value:
        raise ValueError(f"Falta configurar variable/parámetro requerido: {name}")
    return "" if value is None else str(value)


def log_params(logger: logging.Logger, names: List[str]) -> None:
    logger.info("Parámetros recibidos:")
    for name in names:
        value = env_str(name, "")
        logger.info("- %s: %s", name, value if value else "<vacío>")


def load_config() -> Config:
    return Config(
        input_file_path=env_str("INPUT_FILE_PATH", required=True),
        output_dir=env_str("OUTPUT_DIR", required=True),
        output_filename=env_str("OUTPUT_FILENAME", "")
    )


def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def contiene(texto: str, patrones: List[str]) -> bool:
    for patron in patrones:
        if re.search(patron, texto, flags=re.IGNORECASE):
            return True
    return False


def extraer_evidencia(texto: str, patrones: List[str]) -> str:
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            inicio = max(0, m.start() - 90)
            fin = min(len(texto), m.end() + 120)
            return texto[inicio:fin].strip()
    return ""


def leer_archivo(ruta: str, logger: logging.Logger) -> pd.DataFrame:
    path = Path(ruta)

    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {path}")

    logger.info("Leyendo archivo: %s", path)

    if path.suffix.lower() == ".csv":
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            return pd.read_csv(path, encoding="latin-1")

    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path)

    raise ValueError("Formato no soportado. Use CSV, XLSX o XLS.")


def buscar_columna_transcripcion(df: pd.DataFrame) -> str:
    candidatos = [
        "text",
        "texto",
        "transcripcion",
        "transcripción",
        "conversation_text",
        "full_text",
        "transcript",
        "body",
        "frases"
    ]

    cols_normalizadas = {normalizar_texto(c): c for c in df.columns}

    for candidato in candidatos:
        candidato_norm = normalizar_texto(candidato)
        if candidato_norm in cols_normalizadas:
            return cols_normalizadas[candidato_norm]

    for col in df.columns:
        col_norm = normalizar_texto(col)
        if "text" in col_norm or "transcrip" in col_norm or "frase" in col_norm or "phrase" in col_norm:
            return col

    raise ValueError(
        "No se encontró una columna de transcripción. "
        "El archivo debe contener una columna como text, transcripcion, transcript o similar."
    )


def clasificar_motivo(transcripcion: Any) -> Dict[str, str]:
    t = normalizar_texto(transcripcion)

    if not t or len(t) < 20:
        return {
            "Motivo raíz": "No identificado / requiere revisión",
            "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja"
        }

    reglas = [
        {
            "motivo": "Migración a Atlántida HN / nueva app",
            "submotivo": "Credenciales o flujo anterior no funcionan en nueva plataforma",
            "patrones": [
                r"atlantida hn",
                r"nueva app",
                r"nueva aplicacion",
                r"nueva plataforma",
                r"migracion",
                r"migrar",
                r"banca digital nueva",
                r"actualizar la app",
                r"app nueva",
                r"antes ingresaba",
                r"ya no puedo entrar"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe código OTP / SMS / Token",
            "submotivo": "El código de verificación no llega o falla el token",
            "patrones": [
                r"no.*(llega|recibo|manda|envia).*(codigo|otp|token|sms|mensaje|correo)",
                r"(codigo|otp|token|sms).*(no.*llega|no.*recibo|no.*manda|no.*envia)",
                r"codigo de verificacion",
                r"token",
                r"otp",
                r"mensaje de texto",
                r"no me cae el codigo",
                r"no me llega el correo",
                r"correo.*codigo"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Usuario bloqueado / requiere desbloqueo",
            "submotivo": "El usuario está bloqueado o el cliente solicita desbloqueo",
            "patrones": [
                r"bloquead",
                r"desbloque",
                r"usuario bloqueado",
                r"me bloqueo",
                r"se bloqueo",
                r"desbloquear usuario",
                r"cuenta bloqueada"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Actualización de datos / teléfono o correo asociado",
            "submotivo": "Datos registrados no permiten validar identidad o recuperar acceso",
            "patrones": [
                r"actualizar.*(datos|telefono|correo|numero)",
                r"cambiar.*(telefono|correo|numero)",
                r"telefono.*registrado",
                r"correo.*registrado",
                r"numero.*registrado",
                r"ya no tengo.*(numero|telefono|correo)",
                r"otro numero",
                r"otro correo",
                r"datos desactualizados",
                r"actualizacion de datos",
                r"validar identidad"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Contraseña temporal vencida / recuperación incompleta",
            "submotivo": "Clave temporal o enlace de recuperación expiró antes de completar el proceso",
            "patrones": [
                r"temporal",
                r"vencid",
                r"expiro",
                r"expir",
                r"caduc",
                r"enlace.*venc",
                r"clave temporal",
                r"contrasena temporal",
                r"recuperacion incompleta"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Cambio de dispositivo / reinstalación",
            "submotivo": "El cliente cambió celular, reinstaló la app o perdió acceso desde otro equipo",
            "patrones": [
                r"cambie.*(celular|telefono|equipo|dispositivo)",
                r"nuevo celular",
                r"nuevo telefono",
                r"reinstal",
                r"instale de nuevo",
                r"otro dispositivo",
                r"perdi.*celular",
                r"formatee",
                r"desinstale"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Cliente nuevo / primer acceso",
            "submotivo": "Cliente recién afiliado o intentando entrar por primera vez",
            "patrones": [
                r"primera vez",
                r"primer acceso",
                r"cliente nuevo",
                r"recien afiliad",
                r"me afilie",
                r"afiliacion",
                r"activar usuario",
                r"crear usuario",
                r"nunca he entrado",
                r"entrar por primera"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe correo / correo incorrecto",
            "submotivo": "El enlace, correo o notificación de recuperación no llega",
            "patrones": [
                r"correo incorrecto",
                r"correo malo",
                r"no recibe correo",
                r"no me llega.*correo",
                r"email",
                r"mail",
                r"bandeja",
                r"correo electronico"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Error de app/web o autenticación",
            "submotivo": "La app/web muestra error, no permite avanzar o rechaza credenciales",
            "patrones": [
                r"error",
                r"no funciona",
                r"falla",
                r"problema con la app",
                r"problema en la app",
                r"no me deja",
                r"no permite",
                r"rechaza",
                r"autenticacion",
                r"autenticar",
                r"pantalla",
                r"pagina",
                r"web",
                r"aplicacion",
                r"se queda cargando"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Olvidó contraseña / no recuerda credenciales",
            "submotivo": "El cliente manifiesta olvido de clave o usuario",
            "patrones": [
                r"olvide",
                r"olvido",
                r"no recuerdo",
                r"no me acuerdo",
                r"recuperar.*(clave|contrasena|usuario)",
                r"cambiar.*contrasena",
                r"restablecer.*contrasena",
                r"resetear.*contrasena",
                r"clave",
                r"contrasena",
                r"contraseña",
                r"usuario"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Dificultad tecnológica / requiere guía",
            "submotivo": "El cliente necesita acompañamiento paso a paso para completar el proceso",
            "patrones": [
                r"no se como",
                r"ayudeme",
                r"guieme",
                r"paso a paso",
                r"no entiendo",
                r"me cuesta",
                r"no puedo hacerlo",
                r"no manejo"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Consulta operativa / validación general NBDA",
            "submotivo": "La llamada no es claramente cambio de contraseña; es soporte general de banca digital",
            "patrones": [
                r"consulta",
                r"informacion",
                r"ayuda",
                r"soporte",
                r"banca digital",
                r"transferencia",
                r"pago",
                r"saldo",
                r"movimiento",
                r"cuenta"
            ],
            "confianza": "Media"
        }
    ]

    for regla in reglas:
        if contiene(t, regla["patrones"]):
            return {
                "Motivo raíz": regla["motivo"],
                "Submotivo raíz": regla["submotivo"],
                "Evidencia detectada": extraer_evidencia(t, regla["patrones"]),
                "Nivel de confianza": regla["confianza"]
            }

    return {
        "Motivo raíz": "No identificado / requiere revisión",
        "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
        "Evidencia detectada": "",
        "Nivel de confianza": "Baja"
    }


def preparar_base_clasificada(df: pd.DataFrame, col_transcripcion: str) -> pd.DataFrame:
    clasificaciones = df[col_transcripcion].apply(clasificar_motivo).apply(pd.Series)

    base = df.copy()
    base.insert(0, "Motivo raíz", clasificaciones["Motivo raíz"])
    base.insert(1, "Submotivo raíz", clasificaciones["Submotivo raíz"])
    base.insert(2, "Evidencia detectada", clasificaciones["Evidencia detectada"])
    base.insert(3, "Nivel de confianza", clasificaciones["Nivel de confianza"])

    if "Transcripción completa" not in base.columns:
        base["Transcripción completa"] = df[col_transcripcion]

    return base


def preparar_pareto(base: pd.DataFrame) -> pd.DataFrame:
    pareto = (
        base.groupby(["Motivo raíz", "Submotivo raíz"], dropna=False)
        .size()
        .reset_index(name="Llamadas")
        .sort_values("Llamadas", ascending=False)
        .reset_index(drop=True)
    )

    total = int(pareto["Llamadas"].sum())

    if total == 0:
        pareto["%"] = 0
        pareto["% acumulado"] = 0
    else:
        pareto["%"] = pareto["Llamadas"] / total
        pareto["% acumulado"] = pareto["%"].cumsum()

    total_row = pd.DataFrame([{
        "Motivo raíz": "Total general",
        "Submotivo raíz": "",
        "Llamadas": total,
        "%": 1 if total > 0 else 0,
        "% acumulado": 1 if total > 0 else 0
    }])

    return pd.concat([pareto, total_row], ignore_index=True)


def aplicar_formato_excel(output_path: str) -> None:
    wb = load_workbook(output_path)

    rojo = "E31B23"
    verde = "215E4B"
    blanco = "FFFFFF"
    azul_claro = "D9EAF7"
    borde_color = "B7C9E2"

    thin = Side(style="thin", color=borde_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=verde if ws.title == "Base Clasificada" else rojo)
            cell.font = Font(color=blanco, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            if "Transcripción" in header or "text" == header.lower():
                width = 80
            elif "Evidencia" in header:
                width = 70
            elif "Submotivo" in header:
                width = 65
            elif "Motivo" in header:
                width = 45
            else:
                width = min(35, max(12, len(header) + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    if "Pareto" in wb.sheetnames:
        ws = wb["Pareto"]

        for row in range(2, ws.max_row + 1):
            motivo = ws.cell(row=row, column=1).value

            if motivo == "Total general":
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=row, column=col)
                    c.fill = PatternFill("solid", fgColor=rojo)
                    c.font = Font(color=blanco, bold=True)
            else:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=azul_claro)

            ws.cell(row=row, column=4).number_format = "0.00%"
            ws.cell(row=row, column=5).number_format = "0.00%"

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14

    wb.save(output_path)


def generar_excel(config: Config, logger: logging.Logger) -> str:
    df = leer_archivo(config.input_file_path, logger)

    if df.empty:
        raise ValueError("El archivo de entrada no contiene registros.")

    col_transcripcion = buscar_columna_transcripcion(df)
    logger.info("Columna de transcripción detectada: %s", col_transcripcion)

    base = preparar_base_clasificada(df, col_transcripcion)
    pareto = preparar_pareto(base)

    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_filename = str(config.output_filename or "").strip()
    if not output_filename:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Analisis_Profundo_Transcripciones_{fecha}.xlsx"

    if not output_filename.lower().endswith(".xlsx"):
        output_filename = output_filename + ".xlsx"

    output_path = output_dir / output_filename

    logger.info("Generando Excel: %s", output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Base Clasificada", index=False)
        pareto.to_excel(writer, sheet_name="Pareto", index=False)

    aplicar_formato_excel(str(output_path))

    logger.info("Excel generado correctamente: %s", output_path)
    logger.info("Registros clasificados: %s", len(base))
    logger.info("Motivos en Pareto: %s", max(0, len(pareto) - 1))

    return str(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Análisis profundo de transcripciones")
    parser.add_argument("--input-file-path", default=env_str("INPUT_FILE_PATH", ""), help="Ruta completa del CSV/Excel de entrada")
    parser.add_argument("--output-dir", default=env_str("OUTPUT_DIR", ""), help="Carpeta de salida")
    parser.add_argument("--output-filename", default=env_str("OUTPUT_FILENAME", ""), help="Nombre opcional del archivo Excel")
    args = parser.parse_args()

    logger = setup_logger()
    start_time = datetime.now()

    try:
        if args.input_file_path:
            os.environ["INPUT_FILE_PATH"] = args.input_file_path
        if args.output_dir:
            os.environ["OUTPUT_DIR"] = args.output_dir
        if args.output_filename:
            os.environ["OUTPUT_FILENAME"] = args.output_filename

        logger.info("=" * 80)
        logger.info("INICIO ANALISIS PROFUNDO DE TRANSCRIPCIONES")
        log_params(logger, ["INPUT_FILE_PATH", "OUTPUT_DIR", "OUTPUT_FILENAME"])
        logger.info("=" * 80)

        config = load_config()
        output_path = generar_excel(config, logger)

        duration = (datetime.now() - start_time).total_seconds()

        logger.info("=" * 80)
        logger.info("RESUMEN FINAL")
        logger.info("Archivo generado: %s", output_path)
        logger.info("Duración total: %.2f segundos", duration)
        logger.info("=" * 80)

        return 0

    except Exception as exc:
        logger.exception("El proceso terminó con error: %s", exc)
        logger.error(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
