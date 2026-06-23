# -*- coding: utf-8 -*-
"""
Análisis automático de transcripciones NBDA / Cambio de contraseña
Genera un Excel con:
  1. Base Clasificada
  2. Pareto

Diseñado para importarse en PyFlow Manager.

Dependencias:
  pip install pandas openpyxl

Parámetros esperados:
  INPUT_FILE_PATH  = Ruta completa del CSV o Excel de transcripciones
  OUTPUT_DIR       = Carpeta donde se guardará el Excel final
  OUTPUT_FILENAME  = Nombre opcional del archivo final
"""

import os
import re
import sys
import argparse
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PARÁMETROS PARA PYFLOW MANAGER
# ============================================================
PYFLOW_PARAMS = {
    "INPUT_FILE_PATH": {
        "type": "text",
        "label": "Ruta del archivo de transcripciones",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports\GNS_Transcripciones_YYYYMMDD.csv"
    },
    "OUTPUT_DIR": {
        "type": "text",
        "label": "Ruta de salida",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports"
    },
    "OUTPUT_FILENAME": {
        "type": "text",
        "label": "Nombre del archivo final",
        "required": False,
        "default": ""
    }
}


# ============================================================
# UTILIDADES
# ============================================================
def normalizar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def contiene(texto: str, patrones) -> bool:
    return any(re.search(p, texto, flags=re.IGNORECASE) for p in patrones)


def buscar_columna_transcripcion(df: pd.DataFrame) -> str:
    candidatos = [
        "text", "texto", "transcripcion", "transcripción", "conversation_text",
        "full_text", "transcript", "body", "frases"
    ]

    cols_normalizadas = {normalizar_texto(c): c for c in df.columns}

    for c in candidatos:
        if normalizar_texto(c) in cols_normalizadas:
            return cols_normalizadas[normalizar_texto(c)]

    # Si no encuentra nombre exacto, busca columnas que contengan palabras clave
    for col in df.columns:
        col_norm = normalizar_texto(col)
        if any(k in col_norm for k in ["text", "transcrip", "frase", "phrase"]):
            return col

    raise ValueError(
        "No se encontró una columna de transcripción. "
        "El archivo debe contener una columna como: text, transcripcion, transcript o similar."
    )


def leer_archivo(ruta: str) -> pd.DataFrame:
    ruta = Path(ruta)

    if not ruta.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {ruta}")

    if ruta.suffix.lower() == ".csv":
        try:
            return pd.read_csv(ruta, encoding="utf-8-sig")
        except UnicodeDecodeError:
            return pd.read_csv(ruta, encoding="latin-1")

    if ruta.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(ruta)

    raise ValueError("Formato no soportado. Use CSV, XLSX o XLS.")


# ============================================================
# CLASIFICACIÓN PROFUNDA POR MOTIVO RAÍZ
# ============================================================
def clasificar_motivo(transcripcion: str) -> dict:
    t = normalizar_texto(transcripcion)

    if not t or len(t) < 20:
        return {
            "Motivo raíz": "No identificado / requiere revisión",
            "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja"
        }

    reglas = [
        {
            "motivo": "Migración a Atlántida HN / nueva app",
            "submotivo": "Credenciales o flujo anterior no funcionan en nueva plataforma",
            "patrones": [
                r"atlantida hn", r"nueva app", r"nueva aplicacion", r"nueva plataforma",
                r"migracion", r"migrar", r"banca digital nueva", r"actualizar la app",
                r"app nueva", r"antes ingresaba", r"ya no puedo entrar"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe código OTP / SMS / Token",
            "submotivo": "El código de verificación no llega o falla el token",
            "patrones": [
                r"no.*(llega|recibo|manda|envia).*(codigo|otp|token|sms|mensaje|correo)",
                r"(codigo|otp|token|sms).*(no.*llega|no.*recibo|no.*manda|no.*envia)",
                r"codigo de verificacion", r"token", r"otp", r"mensaje de texto",
                r"no me cae el codigo", r"no me llega el correo", r"correo.*codigo"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Error de app/web o autenticación",
            "submotivo": "La app/web muestra error, no permite avanzar o rechaza credenciales",
            "patrones": [
                r"error", r"no funciona", r"falla", r"problema con la app", r"problema en la app",
                r"no me deja", r"no permite", r"rechaza", r"autenticacion", r"autenticar",
                r"pantalla", r"pagina", r"web", r"aplicacion", r"se queda cargando"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Usuario bloqueado / requiere desbloqueo",
            "submotivo": "El usuario está bloqueado o el cliente solicita desbloqueo",
            "patrones": [
                r"bloquead", r"desbloque", r"usuario bloqueado", r"me bloqueo",
                r"se bloqueo", r"desbloquear usuario", r"cuenta bloqueada"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Actualización de datos / teléfono o correo asociado",
            "submotivo": "Datos registrados no permiten validar identidad o recuperar acceso",
            "patrones": [
                r"actualizar.*(datos|telefono|correo|numero)", r"cambiar.*(telefono|correo|numero)",
                r"telefono.*registrado", r"correo.*registrado", r"numero.*registrado",
                r"ya no tengo.*(numero|telefono|correo)", r"otro numero", r"otro correo",
                r"datos desactualizados", r"actualizacion de datos", r"validar identidad"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Olvidó contraseña / no recuerda credenciales",
            "submotivo": "El cliente manifiesta olvido de clave o usuario",
            "patrones": [
                r"olvide", r"olvido", r"no recuerdo", r"no me acuerdo",
                r"recuperar.*(clave|contrasena|usuario)", r"cambiar.*contrasena",
                r"restablecer.*contrasena", r"resetear.*contrasena", r"clave",
                r"contrasena", r"contraseña", r"usuario"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Cliente nuevo / primer acceso",
            "submotivo": "Cliente recién afiliado o intentando entrar por primera vez",
            "patrones": [
                r"primera vez", r"primer acceso", r"cliente nuevo", r"recien afiliad",
                r"me afilie", r"afiliacion", r"activar usuario", r"crear usuario",
                r"nunca he entrado", r"entrar por primera"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Consulta operativa / validación general NBDA",
            "submotivo": "La llamada no es claramente cambio de contraseña; es soporte general de banca digital",
            "patrones": [
                r"consulta", r"informacion", r"ayuda", r"soporte", r"banca digital",
                r"transferencia", r"pago", r"saldo", r"movimiento", r"cuenta"
            ],
            "confianza": "Media"
        },
        {
            "motivo": "Cambio de dispositivo / reinstalación",
            "submotivo": "El cliente cambió celular, reinstaló la app o perdió acceso desde otro equipo",
            "patrones": [
                r"cambie.*(celular|telefono|equipo|dispositivo)", r"nuevo celular",
                r"nuevo telefono", r"reinstal", r"instale de nuevo", r"otro dispositivo",
                r"perdi.*celular", r"formatee", r"desinstale"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Contraseña temporal vencida / recuperación incompleta",
            "submotivo": "Clave temporal o enlace de recuperación expiró antes de completar el proceso",
            "patrones": [
                r"temporal", r"vencid", r"expiro", r"expir", r"caduc",
                r"enlace.*venc", r"clave temporal", r"contrasena temporal",
                r"recuperacion incompleta"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "No recibe correo / correo incorrecto",
            "submotivo": "El enlace, correo o notificación de recuperación no llega",
            "patrones": [
                r"correo incorrecto", r"correo malo", r"no recibe correo",
                r"no me llega.*correo", r"email", r"mail", r"bandeja",
                r"correo electronico"
            ],
            "confianza": "Alta"
        },
        {
            "motivo": "Dificultad tecnológica / requiere guía",
            "submotivo": "El cliente necesita acompañamiento paso a paso para completar el proceso",
            "patrones": [
                r"no se como", r"ayudeme", r"guieme", r"paso a paso",
                r"no entiendo", r"me cuesta", r"no puedo hacerlo", r"no manejo"
            ],
            "confianza": "Media"
        }
    ]

    # Priorización: se asigna el primer motivo fuerte detectado.
    for regla in reglas:
        if contiene(t, regla["patrones"]):
            evidencia = extraer_evidencia(t, regla["patrones"])
            return {
                "Motivo raíz": regla["motivo"],
                "Submotivo raíz": regla["submotivo"],
                "Evidencia detectada": evidencia,
                "Nivel de confianza": regla["confianza"]
            }

    return {
        "Motivo raíz": "No identificado / requiere revisión",
        "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
        "Evidencia detectada": "",
        "Nivel de confianza": "Baja"
    }


def extraer_evidencia(texto: str, patrones) -> str:
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            inicio = max(0, m.start() - 70)
            fin = min(len(texto), m.end() + 90)
            return texto[inicio:fin].strip()
    return ""


# ============================================================
# GENERACIÓN DEL EXCEL
# ============================================================
def preparar_base_clasificada(df: pd.DataFrame, col_transcripcion: str) -> pd.DataFrame:
    clasificaciones = df[col_transcripcion].apply(clasificar_motivo).apply(pd.Series)

    base = df.copy()
    base.insert(0, "Motivo raíz", clasificaciones["Motivo raíz"])
    base.insert(1, "Submotivo raíz", clasificaciones["Submotivo raíz"])
    base.insert(2, "Evidencia detectada", clasificaciones["Evidencia detectada"])
    base.insert(3, "Nivel de confianza", clasificaciones["Nivel de confianza"])

    if "Transcripción completa" not in base.columns:
        base["Transcripción completa"] = df[col_transcripcion]

    return base


def preparar_pareto(base: pd.DataFrame) -> pd.DataFrame:
    pareto = (
        base.groupby(["Motivo raíz", "Submotivo raíz"], dropna=False)
        .size()
        .reset_index(name="Llamadas")
        .sort_values("Llamadas", ascending=False)
        .reset_index(drop=True)
    )

    total = pareto["Llamadas"].sum()
    pareto["%"] = pareto["Llamadas"] / total
    pareto["% acumulado"] = pareto["%"].cumsum()

    total_row = pd.DataFrame([{
        "Motivo raíz": "Total general",
        "Submotivo raíz": "",
        "Llamadas": total,
        "%": 1,
        "% acumulado": 1
    }])

    return pd.concat([pareto, total_row], ignore_index=True)


def aplicar_formato_excel(output_path: str):
    wb = load_workbook(output_path)

    rojo = "E31B23"
    verde = "215E4B"
    blanco = "FFFFFF"
    azul_claro = "D9EAF7"
    borde_color = "B7C9E2"

    thin = Side(style="thin", color=borde_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=verde if ws.title == "Base Clasificada" else rojo)
            cell.font = Font(color=blanco, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            if "Transcripción" in header or "Evidencia" in header:
                width = 70
            elif "Submotivo" in header:
                width = 60
            elif "Motivo" in header:
                width = 45
            else:
                width = min(30, max(12, len(header) + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Formato especial Pareto
    if "Pareto" in wb.sheetnames:
        ws = wb["Pareto"]

        for row in range(2, ws.max_row + 1):
            motivo = ws.cell(row=row, column=1).value

            if motivo == "Total general":
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=row, column=col)
                    c.fill = PatternFill("solid", fgColor=rojo)
                    c.font = Font(color=blanco, bold=True)
            else:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=azul_claro)

            ws.cell(row=row, column=4).number_format = "0.00%"
            ws.cell(row=row, column=5).number_format = "0.00%"

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 75
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14

    wb.save(output_path)


def generar_excel(input_file_path: str, output_dir: str, output_filename: str = "") -> str:
    df = leer_archivo(input_file_path)
    col_transcripcion = buscar_columna_transcripcion(df)

    base = preparar_base_clasificada(df, col_transcripcion)
    pareto = preparar_pareto(base)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not output_filename:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Analisis_Profundo_Transcripciones_{fecha}.xlsx"

    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    output_path = output_dir / output_filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Base Clasificada", index=False)
        pareto.to_excel(writer, sheet_name="Pareto", index=False)

    aplicar_formato_excel(str(output_path))

    return str(output_path)


# ============================================================
# ENTRADA PRINCIPAL
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--INPUT_FILE_PATH", default=os.getenv("INPUT_FILE_PATH", ""))
    parser.add_argument("--OUTPUT_DIR", default=os.getenv("OUTPUT_DIR", ""))
    parser.add_argument("--OUTPUT_FILENAME", default=os.getenv("OUTPUT_FILENAME", ""))

    args = parser.parse_args()

    if not args.INPUT_FILE_PATH:
        raise ValueError("Debe indicar INPUT_FILE_PATH.")

    if not args.OUTPUT_DIR:
        raise ValueError("Debe indicar OUTPUT_DIR.")

    output_path = generar_excel(
        input_file_path=args.INPUT_FILE_PATH,
        output_dir=args.OUTPUT_DIR,
        output_filename=args.OUTPUT_FILENAME
    )

    print(f"Archivo generado correctamente: {output_path}")


if __name__ == "__main__":
    main()
