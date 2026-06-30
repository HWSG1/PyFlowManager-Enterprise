# =========================================================
# GNS Recurrencia AOL / NBDA General
# =========================================================
#
# OBJETIVO
# Calcula recurrencia general por canal:
# - AOL: clientes con mas de una llamada AOL, aunque sean conclusiones distintas.
# - NBDA: clientes con mas de una llamada NBDA, aunque sean conclusiones distintas.
#
# Este script reutiliza la extraccion estable de GNS_Conclusiones_NBDA_AOL.py
# para mantener los mismos filtros de conclusiones AOL/NBDA.
#
# =========================================================

import os
import sys
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from types import SimpleNamespace
from typing import Dict, Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

import GNS_Conclusiones_NBDA_AOL as base


PYFLOW_PARAMS = {
    "GENESYS_CLIENT_ID": {
        "type": "global",
        "global_key": "GENESYS_CLIENT_ID",
        "label": "Genesys Client ID",
        "required": True
    },
    "GENESYS_CLIENT_SECRET": {
        "type": "global",
        "global_key": "GENESYS_CLIENT_SECRET",
        "label": "Genesys Client Secret",
        "required": True,
        "secret": True
    },
    "GENESYS_REGION": {
        "type": "global",
        "global_key": "GENESYS_REGION",
        "label": "Genesys Region / Domain",
        "required": True
    },
    "START_DATE": {
        "type": "date",
        "label": "Fecha inicial local",
        "required": False
    },
    "END_DATE": {
        "type": "date",
        "label": "Fecha final local",
        "required": False
    },
    "AUTO_START_DAY": {
        "type": "number",
        "label": "Dia inicial automatico del mes",
        "required": False,
        "default": "5"
    },
    "GENESYS_TIMEZONE": {
        "type": "text",
        "label": "Zona horaria Genesys",
        "required": True,
        "default": "America/Tegucigalpa"
    },
    "OUTPUT_DIR": {
        "type": "text",
        "label": "Carpeta de salida del Excel",
        "required": False,
        "default": "runtime/exports"
    },
    "DETAIL_PAGE_SIZE": {
        "type": "number",
        "label": "Tamano pagina detalle Genesys",
        "required": False,
        "default": "100"
    }
}


def emit_progress(value: int) -> None:
    value = max(0, min(100, int(value)))
    print(f"PYFLOW_PROGRESS={value}", flush=True)


def env_value(name: str, default: str = "") -> str:
    return str(os.getenv(name, default) or "").strip()


def build_args() -> SimpleNamespace:
    return SimpleNamespace(
        date=None,
        start_date=env_value("START_DATE") or None,
        end_date=env_value("END_DATE") or None,
        start_utc=None,
        end_utc=None,
    )


def build_general_channel_recurrence(df_calls: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    summary_columns = [
        "Canal",
        "Clientes unicos",
        "Llamadas totales",
        "1 consulta",
        "2 consultas",
        "3 consultas",
        "4+ consultas",
        "Clientes recurrentes",
        "% recurrentes",
    ]
    detail_columns = [
        "Canal",
        "externalTag",
        "Llamadas",
        "Conclusiones distintas",
        "Conclusiones",
        "Primera llamada",
        "Ultima llamada",
        "Es recurrente",
    ]

    if df_calls.empty:
        return {
            "summary": pd.DataFrame(columns=summary_columns),
            "detail": pd.DataFrame(columns=detail_columns),
        }

    work = df_calls.copy()
    work["externalTag"] = work["externalTag"].astype(str).str.strip()
    work = work[work["externalTag"].ne("")]
    work = work[work["Banca"].astype(str).str.upper().isin(["AOL", "NBDA"])]

    if work.empty:
        return {
            "summary": pd.DataFrame(columns=summary_columns),
            "detail": pd.DataFrame(columns=detail_columns),
        }

    detail = (
        work.groupby(["Banca", "externalTag"], dropna=False)
        .agg(
            Llamadas=("conversationId", "nunique"),
            Conclusiones_distintas=("Nombre de conclusion", "nunique"),
            Conclusiones=("Nombre de conclusion", lambda values: ", ".join(sorted({str(v) for v in values if str(v).strip()}))),
            Primera_llamada=("conversationStart", "min"),
            Ultima_llamada=("conversationStart", "max"),
        )
        .reset_index()
        .rename(columns={
            "Banca": "Canal",
            "Conclusiones_distintas": "Conclusiones distintas",
            "Primera_llamada": "Primera llamada",
            "Ultima_llamada": "Ultima llamada",
        })
    )
    detail["Es recurrente"] = detail["Llamadas"].astype(int).gt(1).map({True: "Si", False: "No"})
    detail = detail[detail_columns]

    rows = []
    for canal in ["AOL", "NBDA"]:
        channel_detail = detail[detail["Canal"].astype(str).str.upper().eq(canal)]
        channel_calls = work[work["Banca"].astype(str).str.upper().eq(canal)]
        unique_clients = int(channel_detail["externalTag"].nunique()) if not channel_detail.empty else 0
        total_calls = int(channel_calls["conversationId"].nunique()) if not channel_calls.empty else 0
        calls_per_client = pd.to_numeric(channel_detail["Llamadas"], errors="coerce").fillna(0)
        c1 = int(calls_per_client.eq(1).sum())
        c2 = int(calls_per_client.eq(2).sum())
        c3 = int(calls_per_client.eq(3).sum())
        c4 = int(calls_per_client.ge(4).sum())
        recurrent = c2 + c3 + c4
        rows.append({
            "Canal": canal,
            "Clientes unicos": unique_clients,
            "Llamadas totales": total_calls,
            "1 consulta": c1,
            "2 consultas": c2,
            "3 consultas": c3,
            "4+ consultas": c4,
            "Clientes recurrentes": recurrent,
            "% recurrentes": recurrent / unique_clients if unique_clients else 0,
        })

    summary = pd.DataFrame(rows, columns=summary_columns)
    return {"summary": summary, "detail": detail.sort_values(["Canal", "Llamadas"], ascending=[True, False])}


def write_excel(output_path: Path, summary: pd.DataFrame, detail: pd.DataFrame, metadata: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Resumen", index=False)
        detail.to_excel(writer, sheet_name="Detalle Clientes", index=False)
        metadata.to_excel(writer, sheet_name="Parametros", index=False)

        wb = writer.book
        header_fill = "DA282D"
        header_font = "FFFFFF"
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=header_fill)
                cell.font = Font(color=header_font, bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                    if isinstance(cell.value, float) and "recurrentes" in str(ws.cell(row=1, column=cell.column).value).lower():
                        cell.number_format = "0.0%"
                    elif isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def main() -> int:
    logger = base.setup_logger()
    emit_progress(1)
    try:
        logger.info("INICIO RECURRENCIA GENERAL AOL / NBDA")
        config = base.load_config()
        args = build_args()
        start_utc, end_utc, interval_label, start_local, end_local = base.parse_dates(args, config.timezone_name)

        logger.info("Periodo: %s", interval_label)
        logger.info("Inicio UTC: %s", start_utc)
        logger.info("Fin UTC: %s", end_utc)
        emit_progress(8)

        token = base.get_access_token(config, logger)
        emit_progress(15)

        wrapup_catalog = base.get_all_wrapup_codes(config, token, logger)
        emit_progress(25)

        df_calls = base.query_conclusion_call_details(
            config=config,
            token=token,
            start_utc=start_utc,
            end_utc=end_utc,
            wrapup_catalog=wrapup_catalog,
            logger=logger,
        )
        emit_progress(75)

        result = build_general_channel_recurrence(df_calls)
        summary = result["summary"]
        detail = result["detail"]

        metadata = pd.DataFrame([
            {"Parametro": "Periodo", "Valor": interval_label},
            {"Parametro": "Inicio local", "Valor": start_local.strftime("%Y-%m-%d %H:%M:%S")},
            {"Parametro": "Fin local", "Valor": end_local.strftime("%Y-%m-%d %H:%M:%S")},
            {"Parametro": "Inicio UTC", "Valor": start_utc},
            {"Parametro": "Fin UTC", "Valor": end_utc},
            {"Parametro": "Regla AOL", "Valor": "Cliente con mas de una llamada AOL, aunque sean conclusiones distintas"},
            {"Parametro": "Regla NBDA", "Valor": "Cliente con mas de una llamada NBDA, aunque sean conclusiones distintas"},
        ])

        output_dir = Path(config.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(ZoneInfo(config.timezone_name)).strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"Reporte_Recurrencia_AOL_NBDA_General_{timestamp}.xlsx"
        write_excel(output_path, summary, detail, metadata)
        emit_progress(100)

        logger.info("Resumen generado:")
        for _, row in summary.iterrows():
            logger.info(
                "%s | clientes=%s | llamadas=%s | recurrentes=%s | pct=%.2f%%",
                row["Canal"],
                int(row["Clientes unicos"]),
                int(row["Llamadas totales"]),
                int(row["Clientes recurrentes"]),
                float(row["% recurrentes"]) * 100,
            )
        logger.info("Excel generado: %s", output_path)
        logger.info("FIN RECURRENCIA GENERAL AOL / NBDA")
        return 0
    except Exception as exc:
        logger.exception("Error generando recurrencia general AOL/NBDA: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
