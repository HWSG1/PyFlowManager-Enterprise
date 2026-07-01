# =========================================================
# GNS ANALISIS PROFUNDO BLOQUEOS TC/TD V6 - CSV/Excel -> Excel
# =========================================================
#
# Objetivo:
#   Clasificar scripts de bloqueo diario de TC/TD por categoria y comercio,
#   normalizando comercios repetidos para generar un Pareto ejecutivo.
#
# Salida Excel:
#   1) Base Clasificada
#   2) Pareto Categorias
#   3) Pareto Comercios
#
# Diseñado para importarse en PyFlow Manager.
#
# Dependencias:
#   pip install pandas openpyxl
#
# Parametros PyFlow:
#   INPUT_FILE_PATH : Ruta completa del CSV o Excel de bloqueos.
#   OUTPUT_DIR      : Carpeta donde se guardara el Excel final.
#   OUTPUT_FILENAME : Nombre opcional del archivo final.
# =========================================================

import os
import re
import sys
import argparse
import logging
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


PYFLOW_PARAMS = {
    "INPUT_FILE_PATH": {"type": "text", "label": "Ruta completa del archivo CSV/Excel de bloqueos", "required": True},
    "OUTPUT_DIR": {"type": "text", "label": "Ruta de salida del Excel", "required": True},
    "OUTPUT_FILENAME": {"type": "text", "label": "Nombre del archivo Excel final opcional", "required": False},
}

LOGGER_NAME = "gns_analisis_profundo_bloqueos_tc"


@dataclass
class Config:
    input_file_path: str
    output_dir: str
    output_filename: str


def setup_logger() -> logging.Logger:
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S")
    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(formatter)
    logger.addHandler(console)
    return logger


def _clean_env_value(value: Any, default: Optional[str] = None) -> Optional[str]:
    if value is None:
        return default
    text = str(value).strip()
    if text == "" or text.lower() in ("null", "none", "undefined"):
        return default
    return text


def env_str(name: str, default: Optional[str] = None, required: bool = False) -> str:
    value = _clean_env_value(os.getenv(name), default)
    if required and not value:
        raise ValueError(f"Falta configurar variable/parametro requerido: {name}")
    return "" if value is None else str(value)


def log_params(logger: logging.Logger, names: List[str]) -> None:
    logger.info("Parametros recibidos:")
    for name in names:
        value = env_str(name, "")
        logger.info("- %s: %s", name, value if value else "<vacio>")


def load_config() -> Config:
    return Config(
        input_file_path=env_str("INPUT_FILE_PATH", required=True),
        output_dir=env_str("OUTPUT_DIR", required=True),
        output_filename=env_str("OUTPUT_FILENAME", ""),
    )


def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass
    texto = str(valor).lower().replace("\xa0", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.replace("*", " ").replace("/", " ").replace("-", " ")
    texto = re.sub(r"[^a-z0-9$.,\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def contiene(texto: str, patrones: List[str]) -> bool:
    return any(re.search(p, texto, flags=re.IGNORECASE) for p in patrones)


def extraer_evidencia(texto_original: str, patrones: List[str]) -> str:
    texto = normalizar_texto(texto_original)
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            inicio = max(0, m.start() - 90)
            fin = min(len(texto), m.end() + 140)
            return texto[inicio:fin].strip()
    return ""


def leer_archivo(ruta: str, logger: logging.Logger) -> pd.DataFrame:
    path = Path(ruta)
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {path}")

    logger.info("Leyendo archivo: %s", path)
    if path.suffix.lower() == ".csv":
        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                return pd.read_csv(path, encoding=enc)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(path)

    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path)

    raise ValueError("Formato no soportado. Use CSV, XLSX o XLS.")


def buscar_columna_script(df: pd.DataFrame) -> str:
    candidatos = ["script", "descripcion", "descripción", "comentario", "detalle", "observacion", "observación", "texto"]
    cols_normalizadas = {normalizar_texto(c): c for c in df.columns}

    for candidato in candidatos:
        if normalizar_texto(candidato) in cols_normalizadas:
            return cols_normalizadas[normalizar_texto(candidato)]

    for col in df.columns:
        col_norm = normalizar_texto(col)
        if any(x in col_norm for x in ("script", "detalle", "observ", "coment", "descripcion", "texto")):
            return col

    raise ValueError("No se encontro una columna de script/detalle. El archivo debe contener una columna como Script, Detalle u Observacion.")


# =========================================================
# CATALOGO MAESTRO DE COMERCIOS
# =========================================================
# Orden importante: primero aliases especificos y luego los generales.
# No se clasifica por GOOGLE como comercio, porque Google normalmente es
# procesador/canal. El objetivo es llegar al comercio real: TikTok, Free Fire, etc.
CATALOGO_COMERCIOS = [
    {"comercio": "TikTok", "tipo": "Red social / contenido", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["tiktok videos", "tiktok live", "tiktok", "google tiktok"]},
    {"comercio": "Free Fire", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["free fire", "fire fire", "garena", "111dots"]},
    {"comercio": "Candy Crush", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["candy crush", "king com", "king.com"]},
    {"comercio": "Dream League Soccer", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["dream league", "first touch games", "dls"]},
    {"comercio": "Roblox", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["roblox"]},
    {"comercio": "Minecraft", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["minecraft", "minicraft"]},
    {"comercio": "Blood Strike", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["blood strike"]},
    {"comercio": "Supercell", "tipo": "Videojuego", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["supercell", "brawl stars", "clash royale", "clash of clans"]},
    {"comercio": "PlayStation", "tipo": "Videojuego / consola", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["playstation", "psn"]},
    {"comercio": "Xbox", "tipo": "Videojuego / consola", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["xbox", "microsoft xbox"]},
    {"comercio": "Steam", "tipo": "Videojuego / plataforma", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["steam", "valve"]},
    {"comercio": "Epic Games", "tipo": "Videojuego / plataforma", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["epic games", "fortnite"]},
    {"comercio": "Riot Games", "tipo": "Videojuego / plataforma", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["riot games", "league of legends", "valorant"]},
    {"comercio": "Google One", "tipo": "Servicio Google", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["google one"]},
    {"comercio": "YouTube", "tipo": "Streaming / contenido", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["youtube", "yt premium"]},
    {"comercio": "Vix", "tipo": "Streaming", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["vix tv", "vix"]},
    {"comercio": "Prime Video", "tipo": "Streaming", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["prime video", "amazon prime"]},
    {"comercio": "Netflix", "tipo": "Streaming", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["netflix"]},
    {"comercio": "Spotify", "tipo": "Streaming / musica", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["spotify"]},
    {"comercio": "Disney+", "tipo": "Streaming", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["disney plus", "disney+"]},
    {"comercio": "HBO Max", "tipo": "Streaming", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["hbo max", "hbomax", "max.com"]},
    {"comercio": "Facebook / Meta", "tipo": "Red social / publicidad", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["facebook", "facebk", "meta platforms"]},
    {"comercio": "Apple Services", "tipo": "Apple / suscripciones", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["apple.com/bill", "apple com bill", "itunes", "app store", "apple services"]},
    {"comercio": "Tiny4Kiddo", "tipo": "App / suscripcion", "categoria": "Fraude digital / juegos, apps y suscripciones", "aliases": ["tiny4kiddo", "tiny kiddo"]},

    {"comercio": "Temu", "tipo": "Marketplace", "categoria": "E-commerce internacional", "aliases": ["temu"]},
    {"comercio": "AliExpress", "tipo": "Marketplace", "categoria": "E-commerce internacional", "aliases": ["aliexpress", "ali express"]},
    {"comercio": "Amazon", "tipo": "Marketplace", "categoria": "E-commerce internacional", "aliases": ["amazon"]},
    {"comercio": "Shein", "tipo": "Marketplace / moda", "categoria": "E-commerce internacional", "aliases": ["shein"]},
    {"comercio": "eBay", "tipo": "Marketplace", "categoria": "E-commerce internacional", "aliases": ["ebay"]},
    {"comercio": "PayPal", "tipo": "Wallet / pago digital", "categoria": "E-commerce internacional", "aliases": ["paypal", "pay pal"]},
    {"comercio": "Adorama", "tipo": "E-commerce / fotografia", "categoria": "E-commerce internacional", "aliases": ["adorama"]},
    {"comercio": "Quick Box", "tipo": "Courier / casillero", "categoria": "E-commerce internacional", "aliases": ["quick box", "quickbox"]},
    {"comercio": "Shopify", "tipo": "E-commerce", "categoria": "E-commerce internacional", "aliases": ["shopify"]},
    {"comercio": "American Airlines", "tipo": "Aerolínea / viajes", "categoria": "E-commerce internacional", "aliases": ["american airlines"]},
    {"comercio": "Hopper", "tipo": "Viajes", "categoria": "E-commerce internacional", "aliases": ["hopper travel", "hopper"]},
    {"comercio": "Global HP Ecommerce", "tipo": "E-commerce / tecnologia", "categoria": "E-commerce internacional", "aliases": ["global hp ecommerce", "hp ecommerce"]},

    {"comercio": "Tengo", "tipo": "Billetera / recarga", "categoria": "Apuestas / juegos de azar / recargas", "aliases": ["tengo"]},
    {"comercio": "Apostemos", "tipo": "Apuestas", "categoria": "Apuestas / juegos de azar / recargas", "aliases": ["apostemos", "lotelhsa"]},
    {"comercio": "Sherlyluna Kash", "tipo": "Billetera / recarga", "categoria": "Apuestas / juegos de azar / recargas", "aliases": ["sherlyluna kash", "kash"]},

    {"comercio": "Walmart", "tipo": "Comercio local / retail", "categoria": "Comercio local / consumo presencial", "aliases": ["walmart"]},
    {"comercio": "Pricesmart", "tipo": "Comercio local / retail", "categoria": "Comercio local / consumo presencial", "aliases": ["pricesmart", "price smart"]},
    {"comercio": "Farmacia", "tipo": "Comercio local / farmacia", "categoria": "Comercio local / consumo presencial", "aliases": ["farmacia"]},
    {"comercio": "Gasolinera", "tipo": "Comercio local / combustible", "categoria": "Comercio local / consumo presencial", "aliases": ["gasolinera", "puma", "uno", "texaco", "shell"]},
    {"comercio": "FUNTEC", "tipo": "Educacion / tecnologia", "categoria": "Educacion / universidad", "aliases": ["funtec cyberpac", "funtec"]},
    {"comercio": "Universidad", "tipo": "Educacion", "categoria": "Educacion / universidad", "aliases": ["universidad", "unitec", "catolica", "colegio"]},
]


REGLAS_CLASIFICACION = [
    {
        "categoria": "Apuestas / juegos de azar / recargas",
        "subcategoria": "Transacciones asociadas a apuestas, loterias, casinos o billeteras/recargas",
        "patrones": [r"lotelhsa", r"apostemos", r"apuesta", r"casino", r"\bbet\b", r"\btengo\b", r"\bkash\b"],
        "prioridad": 1,
        "confianza": "Alta",
    },
    {
        "categoria": "Patron de multiples transacciones",
        "subcategoria": "Varias transacciones o intentos en poco tiempo",
        "patrones": [r"varias trx", r"\b[2-9]\s*trx", r"\b1[0-9]\s*trx", r"menos de una hora", r"patron fraude", r"patron de fraude", r"multiples transacciones"],
        "prioridad": 2,
        "confianza": "Alta",
    },
    {
        "categoria": "Transacciones denegadas / validacion de seguridad",
        "subcategoria": "Intentos denegados, CVV incorrecto, declined o validacion previa requerida",
        "patrones": [r"denegad", r"declined", r"cvv", r"no match", r"validar", r"verificar", r"antes de activar"],
        "prioridad": 3,
        "confianza": "Media",
    },
    {
        "categoria": "Monto alto o consumo inusual",
        "subcategoria": "Bloqueo por monto elevado o consumo fuera del patron normal",
        "patrones": [r"l\.\s*(?:[1-9][0-9]{3,}|[1-9][0-9]?\.[0-9]{3})", r"lps\s*(?:[1-9][0-9]{3,}|[1-9][0-9]?[,\.][0-9]{3})", r"\$\s*(?:[1-9][0-9]{2,}|[1-9][0-9]?\.[0-9]{3})", r"monto alto", r"sospechoso", r"consumo inusual"],
        "prioridad": 4,
        "confianza": "Media",
    },
    {
        "categoria": "Comercio local / consumo presencial",
        "subcategoria": "Transacciones en comercios locales o consumo nacional",
        "patrones": [r"\bhn\b", r"walmart", r"super", r"farmacia", r"gasolinera", r"restaurante", r"boulevard", r"tegucigalpa", r"cortes", r"francisco morazan", r"san pedro", r"comercial", r"oficina central", r"corven"],
        "prioridad": 5,
        "confianza": "Media",
    },
    {
        "categoria": "E-commerce internacional",
        "subcategoria": "Compras en comercios internacionales o marketplaces",
        "patrones": [r"ecommerce", r"e commerce", r"\.com", r"\b(us|ie|nl|br|mx|es|cn|pa|cr|sv|gt|sg)\b"],
        "prioridad": 6,
        "confianza": "Media",
    },
]


def limpiar_comercio_raw(valor: str) -> str:
    comercio = str(valor or "").replace("\xa0", " ").strip()
    comercio = re.sub(r"\s+", " ", comercio)
    comercio = re.sub(r"^(google|apple|paypal|meta|facebk|facebook)\s*[*\- ]+", "", comercio, flags=re.IGNORECASE)
    comercio = re.sub(r"\b(HN|US|IE|NL|SG|BR|MX|ES|CN|PA|CR|SV|GT)\b.*$", "", comercio, flags=re.IGNORECASE).strip()
    comercio = re.sub(r"\b(aprobadas?|denegadas?|declined|por|lps|l\.|usd|\$)\b.*$", "", comercio, flags=re.IGNORECASE).strip()
    comercio = comercio.strip(" -.,*/")
    return comercio


def extraer_comercio_raw(script: Any) -> str:
    texto = str(script or "").replace("\xa0", " ")
    patrones = [
        r"(?:trx|transaccion|transacciones)\s+(?:en|desde)\s+(.{3,80}?)(?:\s{2,}|\s+por\s+|\s+HN\b|\s+US\b|\s+IE\b|\s+NL\b|\s+SG\b|\s+BR\b|\s+MX\b|\.|$)",
        r"comercio\s+(.{3,80}?)(?:\s{2,}|\s+por\s+|\.|$)",
        r"validar\s+trx\s+en\s+(.{3,80}?)(?:\s{2,}|\s+por\s+|\.|$)",
    ]
    for p in patrones:
        m = re.search(p, texto, flags=re.IGNORECASE)
        if m:
            comercio = limpiar_comercio_raw(m.group(1))
            if comercio:
                return comercio.upper()[:80]
    return "No identificado"


def detectar_comercio_catalogo(script: Any) -> Dict[str, str]:
    t = normalizar_texto(script)
    for item in CATALOGO_COMERCIOS:
        for alias in item["aliases"]:
            alias_norm = normalizar_texto(alias)
            if alias_norm and re.search(re.escape(alias_norm), t):
                return {
                    "Comercio detectado": item["comercio"],
                    "Tipo comercio": item["tipo"],
                    "Categoria catalogo": item["categoria"],
                    "Alias detectado": alias,
                    "Comercio original": extraer_comercio_raw(script),
                }
    raw = extraer_comercio_raw(script)
    if raw != "No identificado":
        return {
            "Comercio detectado": raw.title(),
            "Tipo comercio": "No catalogado",
            "Categoria catalogo": "",
            "Alias detectado": "Sin alias catalogado",
            "Comercio original": raw,
        }
    return {
        "Comercio detectado": "No identificado",
        "Tipo comercio": "No identificado",
        "Categoria catalogo": "",
        "Alias detectado": "Sin alias catalogado",
        "Comercio original": "No identificado",
    }


def obtener_grupo_comercio(tipo_comercio: str, categoria: str = "", comercio: str = "") -> str:
    """Agrupa tipos/comercios a nivel ejecutivo para el Pareto de Comercios.

    La idea es evitar demasiadas filas: TikTok, Facebook y similares se ven como
    Redes sociales; Free Fire, Roblox, Candy Crush, etc. se ven como Videojuegos;
    Netflix, YouTube, Vix, etc. se ven como Streaming, y así sucesivamente.
    """
    tipo = normalizar_texto(tipo_comercio)
    cat = normalizar_texto(categoria)
    com = normalizar_texto(comercio)

    if not tipo or tipo in ("no identificado", "no catalogado"):
        if "script vacio" in cat:
            return "Script vacío / requiere revisión"
        if "no identificado" in cat:
            return "No identificado / requiere revisión"
        if "denegad" in cat or "validacion" in cat:
            return "Validación / transacciones denegadas"
        if "monto alto" in cat or "inusual" in cat:
            return "Monto alto / consumo inusual"
        if "multiple" in cat or "patron" in cat:
            return "Patrón de múltiples transacciones"
        return "Comercio no catalogado"

    if "red social" in tipo or any(x in com for x in ("tiktok", "facebook", "meta", "instagram")):
        return "Redes sociales"
    if "videojuego" in tipo or any(x in com for x in ("free fire", "roblox", "candy crush", "dream league", "minecraft", "supercell", "playstation", "xbox", "steam", "epic games", "riot games")):
        return "Videojuegos"
    if "streaming" in tipo or any(x in com for x in ("netflix", "spotify", "youtube", "vix", "prime video", "disney", "hbo")):
        return "Streaming / contenido digital"
    if "marketplace" in tipo or "e-commerce" in tipo or "ecommerce" in tipo:
        return "E-commerce / marketplaces"
    if "wallet" in tipo or "pago digital" in tipo:
        return "Wallets / pagos digitales"
    if "apuesta" in tipo or "recarga" in tipo or "casino" in tipo:
        return "Apuestas / recargas"
    if "viaje" in tipo or "aerolinea" in tipo or "aerolínea" in tipo:
        return "Viajes / aerolíneas"
    if "courier" in tipo or "casillero" in tipo:
        return "Courier / casilleros"
    if "educacion" in tipo or "universidad" in tipo:
        return "Educación"
    if "comercio local" in tipo or "retail" in tipo or "farmacia" in tipo or "gasolinera" in tipo:
        return "Comercio local / consumo presencial"
    if "apple" in tipo or "google" in tipo or "app" in tipo or "suscripcion" in tipo or "suscripción" in tipo:
        return "Apps / suscripciones"

    return tipo_comercio or "Comercio no catalogado"


def extraer_estado(script: Any) -> str:
    t = normalizar_texto(script)
    aprob = bool(re.search(r"aprobad", t))
    den = bool(re.search(r"denegad|declined|cvv|no match", t))
    if aprob and den:
        return "Aprobadas y denegadas"
    if aprob:
        return "Aprobadas"
    if den:
        return "Denegadas"
    return "No indicado"


def extraer_contactabilidad(script: Any) -> str:
    t = normalizar_texto(script)
    if re.search(r"no contesta|no responde|no se localiza|sin respuesta", t):
        return "Titular no contesta"
    if re.search(r"se llama|se contacto|se contacta|cliente confirma|th confirma", t):
        return "Contacto/llamada mencionada"
    return "No indicado"


def clasificar_por_reglas(texto_original: str) -> Dict[str, str]:
    t = normalizar_texto(texto_original)
    for regla in sorted(REGLAS_CLASIFICACION, key=lambda x: x["prioridad"]):
        if contiene(t, regla["patrones"]):
            return {
                "Categoria": regla["categoria"],
                "Subcategoria": regla["subcategoria"],
                "Regla aplicada": " | ".join(regla["patrones"][:8]),
                "Evidencia detectada": extraer_evidencia(texto_original, regla["patrones"]),
                "Nivel de confianza": regla["confianza"],
            }
    return {
        "Categoria": "No identificado / requiere revision",
        "Subcategoria": "No coincide con reglas actuales",
        "Regla aplicada": "Sin coincidencia",
        "Evidencia detectada": "",
        "Nivel de confianza": "Baja",
    }


def clasificar_script(script: Any) -> Dict[str, str]:
    texto_original = "" if script is None else str(script)
    t = normalizar_texto(texto_original)
    comercio = detectar_comercio_catalogo(texto_original)

    if not t or len(t) < 15:
        return {
            "Categoria": "Script vacio / requiere revision",
            "Subcategoria": "El campo Script viene vacio o sin suficiente detalle",
            "Comercio detectado": "No identificado",
            "Tipo comercio": "No identificado",
            "Grupo comercio": "Script vacío / requiere revisión",
            "Comercio original": "No identificado",
            "Estado transaccion": "No indicado",
            "Contactabilidad": "No indicado",
            "Regla aplicada": "Sin regla",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja",
        }

    regla = clasificar_por_reglas(texto_original)

    # Si el catalogo identifica el comercio, la categoria del catalogo tiene prioridad
    # porque suele ser mas especifica que una regla textual general.
    categoria_final = comercio["Categoria catalogo"] or regla["Categoria"]
    subcategoria_final = regla["Subcategoria"]
    if comercio["Categoria catalogo"]:
        subcategoria_final = f"Comercio catalogado como {comercio['Tipo comercio']}"

    return {
        "Categoria": categoria_final,
        "Subcategoria": subcategoria_final,
        "Comercio detectado": comercio["Comercio detectado"],
        "Tipo comercio": comercio["Tipo comercio"],
        "Grupo comercio": obtener_grupo_comercio(comercio["Tipo comercio"], categoria_final, comercio["Comercio detectado"]),
        "Comercio original": comercio["Comercio original"],
        "Estado transaccion": extraer_estado(texto_original),
        "Contactabilidad": extraer_contactabilidad(texto_original),
        "Regla aplicada": comercio["Alias detectado"] if comercio["Alias detectado"] != "Sin alias catalogado" else regla["Regla aplicada"],
        "Evidencia detectada": regla["Evidencia detectada"],
        "Nivel de confianza": "Alta" if comercio["Categoria catalogo"] else regla["Nivel de confianza"],
    }


def preparar_base_clasificada(df: pd.DataFrame, col_script: str) -> pd.DataFrame:
    clasificaciones = df[col_script].apply(clasificar_script).apply(pd.Series)
    base = df.copy()
    for idx, col in enumerate(clasificaciones.columns):
        base.insert(idx, col, clasificaciones[col])
    if "Script completo" not in base.columns:
        base["Script completo"] = df[col_script]
    return base


def preparar_pareto(base: pd.DataFrame, group_cols: List[str], count_name: str = "Registros") -> pd.DataFrame:
    pareto = (
        base.groupby(group_cols, dropna=False)
        .size()
        .reset_index(name=count_name)
        .sort_values(count_name, ascending=False)
        .reset_index(drop=True)
    )
    total = int(pareto[count_name].sum())
    pareto["%"] = pareto[count_name] / total if total else 0
    pareto["% acumulado"] = pareto["%"].cumsum() if total else 0

    total_row = {c: "" for c in pareto.columns}
    total_row[group_cols[0]] = "Total general"
    total_row[count_name] = total
    total_row["%"] = 1 if total else 0
    total_row["% acumulado"] = 1 if total else 0
    return pd.concat([pareto, pd.DataFrame([total_row])], ignore_index=True)





def preparar_pareto_comercios(base: pd.DataFrame, count_name: str = "Registros") -> pd.DataFrame:
    """Pareto de comercios listo para presentación.

    Muestra una fila por comercio normalizado dentro de su grupo, por ejemplo:
    E-commerce / marketplaces -> Temu 40, AliExpress 9, Amazon 3.

    Para que la hoja no quede demasiado larga, se muestran los comercios más
    relevantes de cada grupo y el resto se consolida en "Otros comercios".
    El detalle completo siempre queda disponible en la hoja Base Clasificada.
    """
    TOP_COMERCIOS_POR_GRUPO = 10

    trabajo = base.copy()
    if "Grupo comercio" not in trabajo.columns:
        trabajo["Grupo comercio"] = trabajo.apply(
            lambda r: obtener_grupo_comercio(
                str(r.get("Tipo comercio", "")),
                str(r.get("Categoria", "")),
                str(r.get("Comercio detectado", "")),
            ),
            axis=1,
        )

    trabajo["Grupo comercio"] = trabajo["Grupo comercio"].fillna("No identificado / requiere revisión").astype(str)
    trabajo["Comercio detectado"] = trabajo["Comercio detectado"].fillna("No identificado").astype(str)
    trabajo["Categoria"] = trabajo["Categoria"].fillna("No identificado / requiere revision").astype(str)

    detalle_base = (
        trabajo.groupby(["Grupo comercio", "Comercio detectado"], dropna=False)
        .agg(
            Registros=("Comercio detectado", "size"),
            Categoria_principal=("Categoria", lambda s: s.value_counts().index[0] if len(s.value_counts()) else ""),
        )
        .reset_index()
    )
    detalle_base["Total grupo"] = detalle_base.groupby("Grupo comercio")["Registros"].transform("sum")

    filas = []
    for grupo, df_grupo in detalle_base.sort_values("Total grupo", ascending=False).groupby("Grupo comercio", sort=False):
        total_grupo = int(df_grupo["Registros"].sum())
        df_grupo = df_grupo.sort_values(["Registros", "Comercio detectado"], ascending=[False, True]).reset_index(drop=True)

        principales = df_grupo.head(TOP_COMERCIOS_POR_GRUPO)
        for _, row in principales.iterrows():
            filas.append({
                "Grupo comercio": grupo,
                "Comercio detectado": row["Comercio detectado"],
                "Registros": int(row["Registros"]),
                "Total grupo": total_grupo,
                "Categoria principal": row["Categoria_principal"],
            })

        otros = df_grupo.iloc[TOP_COMERCIOS_POR_GRUPO:]
        if not otros.empty:
            filas.append({
                "Grupo comercio": grupo,
                "Comercio detectado": "Otros comercios",
                "Registros": int(otros["Registros"].sum()),
                "Total grupo": total_grupo,
                "Categoria principal": otros["Categoria_principal"].value_counts().index[0] if len(otros["Categoria_principal"].value_counts()) else "",
            })

    detalle = pd.DataFrame(filas)
    if detalle.empty:
        detalle = pd.DataFrame(columns=["Grupo comercio", "Comercio detectado", "Registros", "Total grupo", "Categoria principal"])

    detalle = detalle.sort_values(["Total grupo", "Grupo comercio", "Registros", "Comercio detectado"], ascending=[False, True, False, True]).reset_index(drop=True)

    total = int(detalle["Registros"].sum()) if not detalle.empty else 0
    detalle["%"] = detalle["Registros"] / total if total else 0
    detalle["% acumulado"] = detalle["%"].cumsum() if total else 0

    # Orden de columnas pensado para copiar a una presentación.
    detalle = detalle[["Grupo comercio", "Comercio detectado", "Registros", "Total grupo", "Categoria principal", "%", "% acumulado"]]

    total_row = {
        "Grupo comercio": "Total general",
        "Comercio detectado": "",
        "Registros": total,
        "Total grupo": total,
        "Categoria principal": "",
        "%": 1 if total else 0,
        "% acumulado": 1 if total else 0,
    }
    return pd.concat([detalle, pd.DataFrame([total_row])], ignore_index=True)


def combinar_grupos_pareto_comercios(ws, border, fill_color: str, font_color: str) -> None:
    """Combina visualmente el Grupo comercio para que aparezca una sola vez.

    El dataframe mantiene cada comercio en su fila, pero en Excel la columna A
    queda combinada por grupo. Esto mejora la presentación y evita repetir
    'E-commerce / marketplaces' en cada comercio.
    """
    if ws.title != "Pareto Comercios" or ws.max_row < 3:
        return

    start_row = 2
    current_group = ws.cell(row=start_row, column=1).value

    for row in range(3, ws.max_row + 2):
        value = ws.cell(row=row, column=1).value if row <= ws.max_row else None
        is_total = str(value or "") == "Total general"

        if value != current_group or is_total:
            end_row = row - 1
            if current_group and str(current_group) != "Total general" and end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                cell = ws.cell(row=start_row, column=1)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color=font_color, bold=True)
                cell.border = border

                for merged_row in range(start_row + 1, end_row + 1):
                    ws.cell(row=merged_row, column=1).border = border

            start_row = row
            current_group = value

    # Resaltar primer comercio de cada grupo y centrar columnas numéricas.
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "") not in ("", "Total general"):
            ws.cell(row=row, column=2).font = Font(bold=True)
        for col in (3, 4, 6, 7):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

def aplicar_formato_excel(output_path: str) -> None:
    wb = load_workbook(output_path)
    rojo = "E31B23"
    verde = "215E4B"
    blanco = "FFFFFF"
    azul_claro = "D9EAF7"
    borde_color = "B7C9E2"
    thin = Side(style="thin", color=borde_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = verde if ws.title == "Base Clasificada" else rojo

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=header_fill)
            cell.font = Font(color=blanco, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            if header in ("Script", "Script completo"):
                width = 85
            elif header in ("Evidencia detectada", "Subcategoria"):
                width = 60
            elif header in ("Categoria",):
                width = 42
            elif header in ("Comercio detectado", "Comercio original"):
                width = 32
            elif header in ("Tipo comercio", "Grupo comercio"):
                width = 32
            elif header in ("Comercios_principales", "Categoria_principal"):
                width = 55
            else:
                width = min(35, max(12, len(header) + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value or "") == "Total general":
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=row, column=col)
                    c.fill = PatternFill("solid", fgColor=rojo)
                    c.font = Font(color=blanco, bold=True)
            elif ws.title.startswith("Pareto"):
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=azul_claro)
                ws.cell(row=row, column=1).font = Font(bold=True)

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col).value or "")
                if header in ("%", "% acumulado"):
                    ws.cell(row=row, column=col).number_format = "0.00%"
                if header in ("Registros", "Llamadas"):
                    ws.cell(row=row, column=col).number_format = "#,##0"

    if "Pareto Comercios" in wb.sheetnames:
        combinar_grupos_pareto_comercios(wb["Pareto Comercios"], border, azul_claro, "000000")

    wb.save(output_path)


def generar_excel(config: Config, logger: logging.Logger) -> str:
    df = leer_archivo(config.input_file_path, logger)
    if df.empty:
        raise ValueError("El archivo de entrada no contiene registros.")

    col_script = buscar_columna_script(df)
    logger.info("Columna de script detectada: %s", col_script)

    base = preparar_base_clasificada(df, col_script)
    # Pareto ejecutivo: agrupa SOLO por Categoria para evitar filas repetidas
    # por Subcategoria. Ejemplo: todos los casos de Fraude digital / juegos,
    # apps y suscripciones quedan en una sola línea.
    pareto_categoria = preparar_pareto(base, ["Categoria"])
    pareto_comercio = preparar_pareto_comercios(base)

    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_filename = str(config.output_filename or "").strip()
    if not output_filename:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Analisis_Profundo_Bloqueos_TC_V6_{fecha}.xlsx"
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    output_path = output_dir / output_filename
    logger.info("Generando Excel: %s", output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Base Clasificada", index=False)
        pareto_categoria.to_excel(writer, sheet_name="Pareto Categorias", index=False)
        pareto_comercio.to_excel(writer, sheet_name="Pareto Comercios", index=False)

    aplicar_formato_excel(str(output_path))
    logger.info("Excel generado correctamente: %s", output_path)
    logger.info("Registros clasificados: %s", len(base))
    logger.info("Categorias en pareto: %s", max(0, len(pareto_categoria) - 1))
    logger.info("Grupos de comercio en pareto: %s", max(0, len(pareto_comercio) - 1))
    return str(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Analisis profundo de bloqueos TC/TD")
    parser.add_argument("--input-file-path", default=env_str("INPUT_FILE_PATH", ""), help="Ruta completa del CSV/Excel de entrada")
    parser.add_argument("--output-dir", default=env_str("OUTPUT_DIR", ""), help="Carpeta de salida")
    parser.add_argument("--output-filename", default=env_str("OUTPUT_FILENAME", ""), help="Nombre opcional del archivo Excel")
    args = parser.parse_args()

    logger = setup_logger()
    start_time = datetime.now()

    try:
        if args.input_file_path:
            os.environ["INPUT_FILE_PATH"] = args.input_file_path
        if args.output_dir:
            os.environ["OUTPUT_DIR"] = args.output_dir
        if args.output_filename:
            os.environ["OUTPUT_FILENAME"] = args.output_filename

        logger.info("=" * 80)
        logger.info("INICIO ANALISIS PROFUNDO BLOQUEOS TC/TD")
        log_params(logger, ["INPUT_FILE_PATH", "OUTPUT_DIR", "OUTPUT_FILENAME"])
        logger.info("=" * 80)

        config = load_config()
        output_path = generar_excel(config, logger)
        duration = (datetime.now() - start_time).total_seconds()

        logger.info("=" * 80)
        logger.info("RESUMEN FINAL")
        logger.info("Archivo generado: %s", output_path)
        logger.info("Duracion total: %.2f segundos", duration)
        logger.info("=" * 80)
        return 0

    except Exception as exc:
        logger.exception("El proceso termino con error: %s", exc)
        logger.error(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
