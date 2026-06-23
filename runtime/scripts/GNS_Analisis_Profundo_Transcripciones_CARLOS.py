# -*- coding: utf-8 -*-
"""
GNS ANALISIS PROFUNDO DE TRANSCRIPCIONES - VERSION EJECUTIVA PARA CARLOS
CSV/Excel -> Excel con Base Clasificada, Pareto y Resumen Ejecutivo

Objetivo:
  Clasificar transcripciones de llamadas por motivo raíz, con mejor precisión,
  y generar un archivo ejecutivo para revisión gerencial.

Genera:
  1) Base Clasificada
  2) Pareto
  3) Resumen Ejecutivo

Compatible con PyFlow Manager.

Dependencias:
  pip install pandas openpyxl

Parámetros PyFlow:
  INPUT_FILE_PATH : Ruta completa del CSV o Excel de transcripciones.
  OUTPUT_DIR      : Carpeta donde se guardará el Excel final.
  OUTPUT_FILENAME : Nombre opcional del archivo final.

Ejemplo:
  python GNS_Analisis_Profundo_Transcripciones_CARLOS.py ^
    --input-file-path "C:\\PyFlow\\runtime\\exports\\GNS_Transcripciones_20260623_091817.csv" ^
    --output-dir "C:\\PyFlow\\runtime\\exports" ^
    --output-filename "Analisis_Profundo_Transcripciones_Carlos.xlsx"
"""

import os
import re
import sys
import argparse
import logging
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# PARÁMETROS PYFLOW MANAGER
# =========================================================
PYFLOW_PARAMS = {
    "INPUT_FILE_PATH": {
        "type": "text",
        "label": "Ruta completa del archivo CSV/Excel de transcripciones",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports\GNS_Transcripciones_YYYYMMDD_HHMMSS.csv"
    },
    "OUTPUT_DIR": {
        "type": "text",
        "label": "Ruta de salida del Excel",
        "required": True,
        "placeholder": r"C:\PyFlow\runtime\exports"
    },
    "OUTPUT_FILENAME": {
        "type": "text",
        "label": "Nombre del archivo Excel final opcional",
        "required": False,
        "default": ""
    }
}

LOGGER_NAME = "gns_analisis_profundo_transcripciones_carlos"


@dataclass
class Config:
    input_file_path: str
    output_dir: str
    output_filename: str


# =========================================================
# LOGGER Y CONFIGURACIÓN
# =========================================================
def setup_logger() -> logging.Logger:
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        "%Y-%m-%d %H:%M:%S"
    )

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(formatter)
    logger.addHandler(console)

    return logger


def _clean_env_value(value: Any, default: Optional[str] = None) -> Optional[str]:
    if value is None:
        return default
    text = str(value).strip()
    if text == "" or text.lower() in ("null", "none", "undefined"):
        return default
    return text


def env_str(name: str, default: Optional[str] = None, required: bool = False) -> str:
    value = _clean_env_value(os.getenv(name), default)
    if required and not value:
        raise ValueError(f"Falta configurar variable/parámetro requerido: {name}")
    return "" if value is None else str(value)


def log_params(logger: logging.Logger, names: List[str]) -> None:
    logger.info("Parámetros recibidos:")
    for name in names:
        value = env_str(name, "")
        logger.info("- %s: %s", name, value if value else "<vacío>")


def load_config() -> Config:
    return Config(
        input_file_path=env_str("INPUT_FILE_PATH", required=True),
        output_dir=env_str("OUTPUT_DIR", required=True),
        output_filename=env_str("OUTPUT_FILENAME", "")
    )


# =========================================================
# UTILIDADES GENERALES
# =========================================================
def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def buscar_patrones(texto: str, patrones: List[str]) -> List[str]:
    encontrados = []
    for patron in patrones:
        if re.search(patron, texto, flags=re.IGNORECASE):
            encontrados.append(patron)
    return encontrados


def extraer_evidencia(texto: str, patrones: List[str]) -> str:
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            inicio = max(0, m.start() - 90)
            fin = min(len(texto), m.end() + 140)
            return texto[inicio:fin].strip()
    return ""


def leer_archivo(ruta: str, logger: logging.Logger) -> pd.DataFrame:
    path = Path(ruta)

    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {path}")

    logger.info("Leyendo archivo: %s", path)

    if path.suffix.lower() == ".csv":
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            return pd.read_csv(path, encoding="latin-1")

    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path)

    raise ValueError("Formato no soportado. Use CSV, XLSX o XLS. Recuerde incluir la extensión del archivo.")


def buscar_columna_transcripcion(df: pd.DataFrame) -> str:
    candidatos = [
        "text", "texto", "transcripcion", "transcripción", "conversation_text",
        "full_text", "transcript", "body", "frases", "phrases"
    ]

    cols_normalizadas = {normalizar_texto(c): c for c in df.columns}

    for candidato in candidatos:
        candidato_norm = normalizar_texto(candidato)
        if candidato_norm in cols_normalizadas:
            return cols_normalizadas[candidato_norm]

    for col in df.columns:
        col_norm = normalizar_texto(col)
        if "text" in col_norm or "transcrip" in col_norm or "frase" in col_norm or "phrase" in col_norm:
            return col

    raise ValueError(
        "No se encontró una columna de transcripción. "
        "El archivo debe contener una columna como text, transcripcion, transcript o similar."
    )


def buscar_columna(df: pd.DataFrame, candidatos: List[str]) -> Optional[str]:
    cols_normalizadas = {normalizar_texto(c): c for c in df.columns}
    for candidato in candidatos:
        candidato_norm = normalizar_texto(candidato)
        if candidato_norm in cols_normalizadas:
            return cols_normalizadas[candidato_norm]
    return None


def calcular_duracion_segundos(row: pd.Series) -> Optional[float]:
    try:
        inicio = pd.to_datetime(row.get("conversationStart"), utc=True, errors="coerce")
        fin = pd.to_datetime(row.get("conversationEnd"), utc=True, errors="coerce")
        if pd.isna(inicio) or pd.isna(fin):
            return None
        segundos = (fin - inicio).total_seconds()
        if segundos < 0:
            return None
        return round(segundos, 0)
    except Exception:
        return None


def segundos_a_mmss(segundos: Any) -> str:
    if segundos is None:
        return ""
    try:
        if pd.isna(segundos):
            return ""
    except Exception:
        pass
    s = int(round(float(segundos), 0))
    return f"{s // 60:02d}:{s % 60:02d}"


def segundos_a_texto(segundos: Any) -> str:
    if segundos is None:
        return "N/D"
    try:
        if pd.isna(segundos):
            return "N/D"
    except Exception:
        pass
    s = int(round(float(segundos), 0))
    return f"{s // 60}m {s % 60}s"


# =========================================================
# REGLAS DE CLASIFICACIÓN PROFUNDA
# =========================================================
REGLAS = [
    {
        "motivo": "No recibe código OTP / SMS / Token",
        "submotivo": "El código de verificación no llega o falla el token",
        "conclusion_sugerida_crm": "NO_RECIBE_CODIGO_OTP",
        "prioridad": 100,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"no.*(llega|recibo|manda|envia|cae).*(codigo|otp|token|sms|mensaje)",
            r"(codigo|otp|token|sms|mensaje).*(no.*llega|no.*recibo|no.*manda|no.*envia|no.*cae)",
            r"codigo de verificacion", r"no me cae el codigo", r"no me llega el codigo",
            r"no recibo el codigo", r"no me llega el sms", r"no recibo sms",
            r"token.*(no funciona|falla|bloqueado|invalido)"
        ],
        "patrones_medios": [r"\botp\b", r"\btoken\b", r"\bsms\b", r"mensaje de texto", r"codigo"]
    },
    {
        "motivo": "No recibe correo / correo incorrecto",
        "submotivo": "El enlace, correo o notificación de recuperación no llega",
        "conclusion_sugerida_crm": "NO_RECIBE_CORREO_RECUPERACION",
        "prioridad": 95,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"no.*(llega|recibo|manda|envia).*(correo|email|mail)",
            r"(correo|email|mail).*(incorrecto|malo|erroneo|no.*llega|no.*recibo)",
            r"correo incorrecto", r"correo malo", r"correo electronico.*incorrecto",
            r"no me llega el correo", r"no recibo el correo", r"enlace.*correo.*no"
        ],
        "patrones_medios": [r"\bcorreo\b", r"\bemail\b", r"\bmail\b", r"bandeja", r"notificacion"]
    },
    {
        "motivo": "Actualización de datos / teléfono o correo asociado",
        "submotivo": "Datos registrados no permiten validar identidad o recuperar acceso",
        "conclusion_sugerida_crm": "ACTUALIZACION_DATOS_CONTACTO",
        "prioridad": 90,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"actualizar.*(datos|telefono|correo|numero|celular)",
            r"cambiar.*(telefono|correo|numero|celular|datos)",
            r"telefono.*registrado", r"correo.*registrado", r"numero.*registrado",
            r"datos.*(desactualizados|actualizar|actualizados)",
            r"ya no tengo.*(numero|telefono|correo|celular)",
            r"perdi.*(numero|telefono|correo|celular)", r"otro numero", r"otro correo",
            r"validar identidad", r"validacion de identidad"
        ],
        "patrones_medios": [r"actualizacion de datos", r"actualizar", r"correo asociado", r"telefono asociado", r"numero asociado"]
    },
    {
        "motivo": "Usuario bloqueado / requiere desbloqueo",
        "submotivo": "El usuario está bloqueado o el cliente solicita desbloqueo",
        "conclusion_sugerida_crm": "USUARIO_BLOQUEADO_DESBLOQUEO",
        "prioridad": 85,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"usuario bloqueado", r"cuenta bloqueada", r"me bloqueo", r"se bloqueo",
            r"desbloquear usuario", r"solicita desbloqueo", r"necesito desbloquear"
        ],
        "patrones_medios": [r"bloquead", r"desbloque"]
    },
    {
        "motivo": "Contraseña temporal vencida / recuperación incompleta",
        "submotivo": "Clave temporal o enlace de recuperación expiró antes de completar el proceso",
        "conclusion_sugerida_crm": "CONTRASEÑA_TEMPORAL_VENCIDA",
        "prioridad": 80,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"clave temporal", r"contrasena temporal", r"contraseña temporal",
            r"temporal.*(vencid|expiro|expir|caduc)",
            r"enlace.*(vencid|expiro|expir|caduc)", r"recuperacion incompleta"
        ],
        "patrones_medios": [r"temporal", r"vencid", r"expiro", r"expir", r"caduc"]
    },
    {
        "motivo": "Cambio de dispositivo / reinstalación",
        "submotivo": "El cliente cambió celular, reinstaló la app o perdió acceso desde otro equipo",
        "conclusion_sugerida_crm": "CAMBIO_DISPOSITIVO_REINSTALACION",
        "prioridad": 75,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"cambie.*(celular|telefono|equipo|dispositivo)",
            r"cambio.*(celular|telefono|equipo|dispositivo)",
            r"nuevo celular", r"nuevo telefono", r"otro dispositivo", r"reinstal",
            r"instale de nuevo", r"desinstale", r"formatee", r"perdi.*(celular|telefono|equipo)"
        ],
        "patrones_medios": [r"dispositivo", r"celular nuevo", r"telefono nuevo"]
    },
    {
        "motivo": "Cliente nuevo / primer acceso",
        "submotivo": "Cliente recién afiliado o intentando entrar por primera vez",
        "conclusion_sugerida_crm": "CLIENTE_NUEVO_PRIMER_ACCESO",
        "prioridad": 70,
        "confianza_base": "Alta",
        "patrones_fuertes": [
            r"primera vez", r"primer acceso", r"cliente nuevo", r"recien afiliad",
            r"me afilie", r"afiliacion", r"activar usuario", r"crear usuario",
            r"nunca he entrado", r"entrar por primera"
        ],
        "patrones_medios": [r"afiliar", r"registro", r"registrarme"]
    },
    {
        "motivo": "Error de app/web o autenticación",
        "submotivo": "La app/web muestra error, no permite avanzar o rechaza credenciales",
        "conclusion_sugerida_crm": "PAGINA_WEB_APP_NO_FUNCIONA",
        "prioridad": 65,
        "confianza_base": "Media",
        "patrones_fuertes": [
            r"app.*(error|falla|no funciona|no abre|no me deja|no permite)",
            r"aplicacion.*(error|falla|no funciona|no abre|no me deja|no permite)",
            r"pagina.*(error|falla|no funciona|no abre)",
            r"web.*(error|falla|no funciona)", r"no me deja entrar",
            r"no permite avanzar", r"rechaza.*credencial", r"credenciales.*incorrect",
            r"autenticacion.*(falla|error|rechaza)"
        ],
        "patrones_medios": [
            r"error", r"no funciona", r"falla", r"no me deja", r"no permite", r"rechaza",
            r"autenticacion", r"autenticar", r"pantalla", r"pagina", r"web", r"aplicacion", r"se queda cargando"
        ]
    },
    {
        "motivo": "Olvidó contraseña / no recuerda credenciales",
        "submotivo": "El cliente manifiesta olvido de clave o usuario",
        "conclusion_sugerida_crm": "OLVIDO_CONTRASEÑA_CREDENCIALES",
        "prioridad": 60,
        "confianza_base": "Media",
        "patrones_fuertes": [
            r"olvide.*(clave|contrasena|contraseña|usuario)",
            r"olvido.*(clave|contrasena|contraseña|usuario)",
            r"no recuerdo.*(clave|contrasena|contraseña|usuario)",
            r"no me acuerdo.*(clave|contrasena|contraseña|usuario)",
            r"recuperar.*(clave|contrasena|contraseña|usuario)",
            r"restablecer.*(clave|contrasena|contraseña)",
            r"resetear.*(clave|contrasena|contraseña)"
        ],
        "patrones_medios": [r"olvide", r"olvido", r"no recuerdo", r"no me acuerdo", r"clave", r"contrasena", r"contraseña", r"usuario"]
    },
    {
        "motivo": "Dificultad tecnológica / requiere guía",
        "submotivo": "El cliente necesita acompañamiento paso a paso para completar el proceso",
        "conclusion_sugerida_crm": "DIFICULTAD_TECNOLOGICA_CLIENTE",
        "prioridad": 55,
        "confianza_base": "Media",
        "patrones_fuertes": [
            r"no se como", r"ayudeme.*paso", r"guieme", r"paso a paso",
            r"no entiendo", r"me cuesta", r"no puedo hacerlo", r"no manejo"
        ],
        "patrones_medios": [r"ayudeme", r"ayuda", r"explicarme"]
    },
    {
        "motivo": "Migración a Atlántida HN / nueva app",
        "submotivo": "Credenciales o flujo anterior no funcionan en nueva plataforma",
        "conclusion_sugerida_crm": "SOPORTE_MIGRACIÓN_ATLANTIDA_HN",
        "prioridad": 50,
        "confianza_base": "Media",
        "patrones_fuertes": [
            r"atlantida hn", r"nueva app", r"nueva aplicacion", r"nueva plataforma",
            r"migracion", r"migrar", r"banca digital nueva", r"app nueva",
            r"antes ingresaba", r"ya no puedo entrar.*(nueva|app|atlantida)"
        ],
        "patrones_medios": [r"actualizar la app", r"banca digital", r"atlantida"]
    },
    {
        "motivo": "Problemas de conectividad del cliente",
        "submotivo": "La falla parece estar asociada a internet, señal, datos móviles, WiFi o equipo del cliente",
        "conclusion_sugerida_crm": "PROBLEMAS_DE_CONECTIVIDAD_DEL_CLIENTE",
        "prioridad": 48,
        "confianza_base": "Media",
        "patrones_fuertes": [
            r"sin internet", r"no tengo internet", r"datos moviles", r"wifi", r"wi fi",
            r"senal", r"señal", r"mala conexion", r"problema de conexion", r"conexion lenta",
            r"red.*(lenta|mala|fallando)", r"internet.*(lento|fallando|malo)"
        ],
        "patrones_medios": [r"internet", r"conexion", r"red", r"senal", r"wifi", r"datos"]
    },
    {
        "motivo": "Consulta operativa / validación general NBDA",
        "submotivo": "La llamada no es claramente cambio de contraseña; es soporte general de banca digital",
        "conclusion_sugerida_crm": "SOPORTE_GENERAL_NBDA",
        "prioridad": 20,
        "confianza_base": "Media",
        "patrones_fuertes": [r"consulta", r"informacion", r"transferencia", r"pago", r"saldo", r"movimiento", r"cuenta"],
        "patrones_medios": [r"soporte", r"banca digital", r"ayuda"]
    }
]


def clasificar_motivo(transcripcion: Any) -> Dict[str, Any]:
    t = normalizar_texto(transcripcion)

    if not t or len(t) < 20:
        return {
            "Motivo raíz": "No identificado / requiere revisión",
            "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
            "Conclusión sugerida CRM": "REQUIERE_REVISION_MANUAL",
            "Motivos secundarios": "",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja",
            "Puntaje clasificación": 0
        }

    candidatos = []

    for regla in REGLAS:
        fuertes = buscar_patrones(t, regla["patrones_fuertes"])
        medios = buscar_patrones(t, regla["patrones_medios"])

        if not fuertes and not medios:
            continue

        puntaje = (len(fuertes) * 10) + (len(medios) * 3) + regla["prioridad"]

        # Evita que migración gane solo por contexto cuando hay un motivo operativo más específico.
        if regla["motivo"] == "Migración a Atlántida HN / nueva app" and not fuertes:
            puntaje -= 20

        # Consulta general solo debe ganar si no hay un motivo más claro.
        if regla["motivo"] == "Consulta operativa / validación general NBDA":
            puntaje -= 25

        candidatos.append({
            "regla": regla,
            "fuertes": fuertes,
            "medios": medios,
            "puntaje": puntaje
        })

    if not candidatos:
        return {
            "Motivo raíz": "No identificado / requiere revisión",
            "Submotivo raíz": "La transcripción no contiene señales suficientes para inferir motivo raíz",
            "Conclusión sugerida CRM": "REQUIERE_REVISION_MANUAL",
            "Motivos secundarios": "",
            "Evidencia detectada": "",
            "Nivel de confianza": "Baja",
            "Puntaje clasificación": 0
        }

    candidatos = sorted(candidatos, key=lambda x: x["puntaje"], reverse=True)
    ganador = candidatos[0]
    regla = ganador["regla"]
    patrones_evidencia = ganador["fuertes"] or ganador["medios"]

    if ganador["fuertes"]:
        confianza = regla["confianza_base"]
    elif regla["confianza_base"] == "Alta":
        confianza = "Media"
    else:
        confianza = "Media"

    if len(candidatos) > 1 and (ganador["puntaje"] - candidatos[1]["puntaje"]) <= 5:
        confianza = "Media"

    secundarios = []
    for candidato in candidatos[1:4]:
        if candidato["puntaje"] >= 50:
            secundarios.append(candidato["regla"]["motivo"])

    return {
        "Motivo raíz": regla["motivo"],
        "Submotivo raíz": regla["submotivo"],
        "Conclusión sugerida CRM": regla["conclusion_sugerida_crm"],
        "Motivos secundarios": " | ".join(secundarios),
        "Evidencia detectada": extraer_evidencia(t, patrones_evidencia),
        "Nivel de confianza": confianza,
        "Puntaje clasificación": ganador["puntaje"]
    }


# =========================================================
# PREPARACIÓN DE HOJAS
# =========================================================
def preparar_base_clasificada(df: pd.DataFrame, col_transcripcion: str) -> pd.DataFrame:
    df = df.copy()
    clasificaciones = df[col_transcripcion].apply(clasificar_motivo).apply(pd.Series)

    if "conversationStart" in df.columns and "conversationEnd" in df.columns:
        df["Duracion_seg"] = df.apply(calcular_duracion_segundos, axis=1)
        df["Duración (MM:SS)"] = df["Duracion_seg"].apply(segundos_a_mmss)

    base = df.copy()

    columnas_insertar = [
        "Motivo raíz",
        "Submotivo raíz",
        "Conclusión sugerida CRM",
        "Motivos secundarios",
        "Evidencia detectada",
        "Nivel de confianza",
        "Puntaje clasificación"
    ]

    for idx, col in enumerate(columnas_insertar):
        base.insert(idx, col, clasificaciones[col])

    if "Transcripción completa" not in base.columns:
        base["Transcripción completa"] = df[col_transcripcion]

    # Colocar transcripción al final, sin borrar la columna original.
    cols = [c for c in base.columns if c != "Transcripción completa"] + ["Transcripción completa"]
    base = base[cols]

    return base


def preparar_pareto(base: pd.DataFrame) -> pd.DataFrame:
    pareto = (
        base.groupby(["Motivo raíz", "Submotivo raíz", "Conclusión sugerida CRM"], dropna=False)
        .agg(
            Llamadas=("Motivo raíz", "size"),
            Duracion_promedio_seg=("Duracion_seg", "mean") if "Duracion_seg" in base.columns else ("Motivo raíz", "size"),
            Confianza_alta=("Nivel de confianza", lambda s: int((s == "Alta").sum())),
            Confianza_media=("Nivel de confianza", lambda s: int((s == "Media").sum())),
            Confianza_baja=("Nivel de confianza", lambda s: int((s == "Baja").sum()))
        )
        .reset_index()
        .sort_values("Llamadas", ascending=False)
        .reset_index(drop=True)
    )

    if "Duracion_promedio_seg" in pareto.columns:
        pareto["Duración promedio"] = pareto["Duracion_promedio_seg"].apply(segundos_a_texto)
        pareto = pareto.drop(columns=["Duracion_promedio_seg"])

    total = int(pareto["Llamadas"].sum())
    if total == 0:
        pareto["%"] = 0
        pareto["% acumulado"] = 0
    else:
        pareto["%"] = pareto["Llamadas"] / total
        pareto["% acumulado"] = pareto["%"].cumsum()

    total_row = {
        "Motivo raíz": "Total general",
        "Submotivo raíz": "",
        "Conclusión sugerida CRM": "",
        "Llamadas": total,
        "Confianza_alta": int((base["Nivel de confianza"] == "Alta").sum()),
        "Confianza_media": int((base["Nivel de confianza"] == "Media").sum()),
        "Confianza_baja": int((base["Nivel de confianza"] == "Baja").sum()),
        "Duración promedio": segundos_a_texto(base["Duracion_seg"].mean()) if "Duracion_seg" in base.columns else "N/D",
        "%": 1 if total > 0 else 0,
        "% acumulado": 1 if total > 0 else 0
    }

    return pd.concat([pareto, pd.DataFrame([total_row])], ignore_index=True)


def preparar_resumen_ejecutivo(df_original: pd.DataFrame, base: pd.DataFrame, pareto: pd.DataFrame) -> List[List[Any]]:
    total = len(base)

    col_dir = buscar_columna(df_original, ["originatingDirection", "direccion", "direction"])
    if col_dir:
        dir_norm = df_original[col_dir].astype(str).str.lower()
        inbound = int((dir_norm == "inbound").sum())
        outbound = int((dir_norm == "outbound").sum())
    else:
        inbound = "N/D"
        outbound = "N/D"

    if "Duracion_seg" in base.columns:
        duraciones = pd.to_numeric(base["Duracion_seg"], errors="coerce").dropna()
        dur_prom = segundos_a_texto(duraciones.mean()) if len(duraciones) else "N/D"
        dur_max = segundos_a_texto(duraciones.max()) if len(duraciones) else "N/D"
        dur_min = segundos_a_texto(duraciones.min()) if len(duraciones) else "N/D"
    else:
        dur_prom = dur_max = dur_min = "N/D"

    conf_alta = int((base["Nivel de confianza"] == "Alta").sum())
    conf_media = int((base["Nivel de confianza"] == "Media").sum())
    conf_baja = int((base["Nivel de confianza"] == "Baja").sum())

    pareto_sin_total = pareto[pareto["Motivo raíz"] != "Total general"].copy()
    if not pareto_sin_total.empty:
        motivo_top = pareto_sin_total.iloc[0]
        top_motivo = motivo_top["Motivo raíz"]
        top_llamadas = int(motivo_top["Llamadas"])
        top_pct = f"{float(motivo_top['%']):.1%}"
    else:
        top_motivo = "N/D"
        top_llamadas = 0
        top_pct = "N/D"

    col_fecha = buscar_columna(df_original, ["conversationStart", "fecha", "fecha_inicio", "conversation_start"])
    if col_fecha:
        fechas = pd.to_datetime(df_original[col_fecha], errors="coerce")
        fecha_inicio = fechas.min()
        fecha_fin = fechas.max()
        fecha_inicio_txt = str(fecha_inicio.date()) if not pd.isna(fecha_inicio) else "N/D"
        fecha_fin_txt = str(fecha_fin.date()) if not pd.isna(fecha_fin) else "N/D"
    else:
        fecha_inicio_txt = "N/D"
        fecha_fin_txt = "N/D"

    no_identificados = int((base["Motivo raíz"] == "No identificado / requiere revisión").sum())

    rows = [
        ["RESUMEN EJECUTIVO", ""],
        ["", ""],
        ["Período analizado", ""],
        ["Fecha más antigua", fecha_inicio_txt],
        ["Fecha más reciente", fecha_fin_txt],
        ["", ""],
        ["Volumen de llamadas", ""],
        ["Total llamadas", total],
        ["Inbound", inbound],
        ["Outbound", outbound],
        ["", ""],
        ["Duración de llamadas", ""],
        ["Promedio", dur_prom],
        ["Máximo", dur_max],
        ["Mínimo", dur_min],
        ["", ""],
        ["Clasificación", ""],
        ["Motivo raíz más frecuente", top_motivo],
        ["Llamadas en motivo top", top_llamadas],
        ["% del total", top_pct],
        ["No identificados / revisión manual", no_identificados],
        ["", ""],
        ["Nivel de confianza", ""],
        ["Alta", conf_alta],
        ["Media", conf_media],
        ["Baja", conf_baja],
        ["", ""],
        ["Lectura ejecutiva", ""],
        ["Mensaje clave", generar_mensaje_ejecutivo(pareto_sin_total, total)]
    ]

    return rows


def generar_mensaje_ejecutivo(pareto_sin_total: pd.DataFrame, total: int) -> str:
    if total == 0 or pareto_sin_total.empty:
        return "No se identificaron registros para análisis."

    top = pareto_sin_total.head(3)
    partes = []
    for _, row in top.iterrows():
        partes.append(f"{row['Motivo raíz']} ({float(row['%']):.1%})")

    return "Los principales motivos detectados fueron: " + "; ".join(partes) + "."


# =========================================================
# FORMATO EXCEL
# =========================================================
VERDE_HEADER = "215E4B"
ROJO_HEADER = "E31B23"
AZUL_HEADER = "1F4E79"
BLANCO = "FFFFFF"
AZUL_CLARO = "D9EAF7"
GRIS_CLARO = "F2F2F2"
BORDE_COLOR = "B7C9E2"
AMARILLO = "FFF2CC"


def _borde() -> Border:
    thin = Side(style="thin", color=BORDE_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _header_font() -> Font:
    return Font(color=BLANCO, bold=True, name="Arial", size=10)


def _data_font(bold: bool = False) -> Font:
    return Font(name="Arial", size=9, bold=bold)


def formato_base_clasificada(ws) -> None:
    border = _borde()

    for cell in ws[1]:
        cell.fill = _header_fill(VERDE_HEADER)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill_row = PatternFill("solid", fgColor=GRIS_CLARO) if i % 2 == 0 else None
        for cell in row:
            cell.border = border
            cell.font = _data_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill_row:
                cell.fill = fill_row

    anchos = {
        "Motivo raíz": 42,
        "Submotivo raíz": 60,
        "Conclusión sugerida CRM": 34,
        "Motivos secundarios": 55,
        "Evidencia detectada": 75,
        "Nivel de confianza": 18,
        "Puntaje clasificación": 18,
        "conversationId": 38,
        "communicationId": 38,
        "conversationStart": 22,
        "conversationEnd": 22,
        "originatingDirection": 18,
        "phrases_count": 14,
        "Duracion_seg": 14,
        "Duración (MM:SS)": 16,
        "fecha_carga": 22,
        "text": 90,
        "Transcripción completa": 90,
    }

    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        width = anchos.get(header, min(35, max(12, len(header) + 4)))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def formato_pareto(ws) -> None:
    border = _borde()

    for cell in ws[1]:
        cell.fill = _header_fill(ROJO_HEADER)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        motivo = ws.cell(row=row_num, column=1).value

        if motivo == "Total general":
            for cell in row:
                cell.fill = _header_fill(ROJO_HEADER)
                cell.font = Font(color=BLANCO, bold=True, name="Arial", size=9)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            for cell in row:
                cell.border = border
                cell.font = _data_font()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row[0].font = Font(bold=True, name="Arial", size=9)
            row[0].fill = _header_fill(AZUL_CLARO)

        # Porcentajes: últimas dos columnas
        ws.cell(row=row_num, column=ws.max_column - 1).number_format = "0.00%"
        ws.cell(row=row_num, column=ws.max_column).number_format = "0.00%"

    widths = {
        "A": 45, "B": 75, "C": 34, "D": 14, "E": 16,
        "F": 16, "G": 16, "H": 18, "I": 12, "J": 16
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def formato_resumen(ws) -> None:
    border = _borde()

    secciones = {
        "Período analizado",
        "Volumen de llamadas",
        "Duración de llamadas",
        "Clasificación",
        "Nivel de confianza",
        "Lectura ejecutiva"
    }

    for row in range(1, ws.max_row + 1):
        etiqueta = ws.cell(row=row, column=1).value
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)

        if etiqueta == "RESUMEN EJECUTIVO":
            for c in (c1, c2):
                c.fill = _header_fill(AZUL_HEADER)
                c.font = Font(color=BLANCO, bold=True, name="Arial", size=13)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border
            ws.row_dimensions[row].height = 26
        elif etiqueta in secciones:
            for c in (c1, c2):
                c.fill = _header_fill(VERDE_HEADER)
                c.font = Font(color=BLANCO, bold=True, name="Arial", size=10)
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = border
        elif etiqueta in (None, ""):
            continue
        else:
            c1.font = _data_font(bold=True)
            c2.font = _data_font()
            c1.border = border
            c2.border = border
            c1.alignment = Alignment(vertical="center", wrap_text=True)
            c2.alignment = Alignment(vertical="center", wrap_text=True)

            if etiqueta == "Mensaje clave":
                c1.fill = _header_fill(AMARILLO)
                c2.fill = _header_fill(AMARILLO)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"


def aplicar_formato_excel(output_path: str) -> None:
    wb = load_workbook(output_path)

    if "Base Clasificada" in wb.sheetnames:
        formato_base_clasificada(wb["Base Clasificada"])

    if "Pareto" in wb.sheetnames:
        formato_pareto(wb["Pareto"])

    if "Resumen Ejecutivo" in wb.sheetnames:
        formato_resumen(wb["Resumen Ejecutivo"])

    wb.save(output_path)


# =========================================================
# PROCESO PRINCIPAL
# =========================================================
def generar_excel(config: Config, logger: logging.Logger) -> str:
    df = leer_archivo(config.input_file_path, logger)

    if df.empty:
        raise ValueError("El archivo de entrada no contiene registros.")

    logger.info("Registros cargados: %s", len(df))
    logger.info("Columnas detectadas: %s", list(df.columns))

    col_transcripcion = buscar_columna_transcripcion(df)
    logger.info("Columna de transcripción detectada: %s", col_transcripcion)

    logger.info("Clasificando transcripciones...")
    base = preparar_base_clasificada(df, col_transcripcion)

    logger.info("Generando Pareto...")
    pareto = preparar_pareto(base)

    logger.info("Generando Resumen Ejecutivo...")
    resumen_rows = preparar_resumen_ejecutivo(df, base, pareto)
    resumen_df = pd.DataFrame(resumen_rows, columns=["Métrica", "Valor"])

    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_filename = str(config.output_filename or "").strip()
    if not output_filename:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Analisis_Profundo_Transcripciones_Carlos_{fecha}.xlsx"

    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    output_path = output_dir / output_filename

    logger.info("Escribiendo Excel: %s", output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        resumen_df.to_excel(writer, sheet_name="Resumen Ejecutivo", index=False, header=False)
        pareto.to_excel(writer, sheet_name="Pareto", index=False)
        base.to_excel(writer, sheet_name="Base Clasificada", index=False)

    logger.info("Aplicando formato ejecutivo...")
    aplicar_formato_excel(str(output_path))

    logger.info("Excel generado correctamente: %s", output_path)
    logger.info("Registros clasificados: %s", len(base))
    logger.info("Motivos en Pareto: %s", max(0, len(pareto) - 1))

    return str(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Análisis profundo de transcripciones - versión ejecutiva para Carlos")
    parser.add_argument("--input-file-path", "--INPUT_FILE_PATH", dest="input_file_path", default=env_str("INPUT_FILE_PATH", ""), help="Ruta completa del CSV/Excel de entrada")
    parser.add_argument("--output-dir", "--OUTPUT_DIR", dest="output_dir", default=env_str("OUTPUT_DIR", ""), help="Carpeta de salida")
    parser.add_argument("--output-filename", "--OUTPUT_FILENAME", dest="output_filename", default=env_str("OUTPUT_FILENAME", ""), help="Nombre opcional del archivo Excel")
    args = parser.parse_args()

    logger = setup_logger()
    start_time = datetime.now()

    try:
        if args.input_file_path:
            os.environ["INPUT_FILE_PATH"] = args.input_file_path
        if args.output_dir:
            os.environ["OUTPUT_DIR"] = args.output_dir
        if args.output_filename:
            os.environ["OUTPUT_FILENAME"] = args.output_filename

        logger.info("=" * 80)
        logger.info("INICIO ANALISIS PROFUNDO DE TRANSCRIPCIONES - VERSION CARLOS")
        log_params(logger, ["INPUT_FILE_PATH", "OUTPUT_DIR", "OUTPUT_FILENAME"])
        logger.info("=" * 80)

        config = load_config()
        output_path = generar_excel(config, logger)

        duration = (datetime.now() - start_time).total_seconds()

        logger.info("=" * 80)
        logger.info("RESUMEN FINAL")
        logger.info("Archivo generado: %s", output_path)
        logger.info("Duración total: %.2f segundos", duration)
        logger.info("=" * 80)

        return 0

    except Exception as exc:
        logger.exception("El proceso terminó con error: %s", exc)
        logger.error(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
